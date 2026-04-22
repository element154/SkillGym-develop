"""Verification tests for the SiO2 mp-6930 Excel report task.

These tests recompute expected values from:
- /root/1272701
- /root/SiO2-Quartz-alpha.cif
- /root/SiO2-Cristobalite.cif

The tests are intentionally strict on workbook layout/formatting, but robust to:
- HTML facts being present in visible text or meta tags
- CIF species carrying oxidation states such as Si4+ / O2-
"""

from __future__ import annotations

import html as html_lib
import re
from pathlib import Path

import numpy as np
import pytest
from openpyxl import load_workbook
from pymatgen.analysis.local_env import CrystalNN
from pymatgen.core import Structure
from pymatgen.symmetry.analyzer import SpacegroupAnalyzer


OUTPUT_FILE = Path("/root/sio2_mp6930_verification.xlsx")
HTML_FILE = Path("/root/1272701")
QUARTZ_CIF = Path("/root/SiO2-Quartz-alpha.cif")
CRISTOBALITE_CIF = Path("/root/SiO2-Cristobalite.cif")
SHEET_NAME = "Verification"

EXPECTED_HEADERS = [
    "Metric",
    "MP mp-6930 (HTML)",
    "Quartz (CIF)",
    "Cristobalite (CIF)",
    "Quartz matches MP?",
]
EXPECTED_METRICS = [
    "Material ID",
    "Space group (symbol)",
    "Space group (number)",
    "Si\u2013O short (\u00c5)",
    "Si\u2013O long (\u00c5)",
    "Si\u2013O\u2013Si angle (deg)",
]

BLUE_RGB = "0000FF"
BLACK_RGB = "000000"
SUBSCRIPT_TRANS = str.maketrans("\u2080\u2081\u2082\u2083\u2084\u2085\u2086\u2087\u2088\u2089", "0123456789")


def _normalize_ws(text: str) -> str:
    return " ".join(str(text).replace("\xa0", " ").split())


def _normalize_symbol(symbol: str) -> str:
    value = str(symbol).translate(SUBSCRIPT_TRANS)
    value = value.replace(" ", "").replace("_", "").replace("-", "")
    return value.upper()


def _element_symbol(site) -> str:
    specie = getattr(site, "specie", None)
    symbol = getattr(specie, "symbol", None)
    if symbol:
        return str(symbol)
    species_string = getattr(site, "species_string", "")
    if species_string:
        match = re.match(r"[A-Z][a-z]?", species_string)
        if match:
            return match.group(0)
    text = str(specie if specie is not None else site)
    match = re.match(r"[A-Z][a-z]?", text)
    return match.group(0) if match else text


def _font_is_blue(cell) -> bool:
    color = cell.font.color
    if color is None:
        return False
    rgb = getattr(color, "rgb", None)
    if rgb is None:
        return False
    return BLUE_RGB in str(rgb).upper()


def _font_is_black_or_default(cell) -> bool:
    color = cell.font.color
    if color is None:
        return True
    rgb = getattr(color, "rgb", None)
    if rgb is None:
        return True
    rgb = str(rgb).upper()
    return (BLACK_RGB in rgb) or (rgb == "00000000")


def _angle_deg(a: np.ndarray, o: np.ndarray, b: np.ndarray) -> float:
    v1 = a - o
    v2 = b - o
    denom = np.linalg.norm(v1) * np.linalg.norm(v2)
    assert denom > 0, "Zero-length vector encountered while computing angle"
    cosang = np.dot(v1, v2) / denom
    cosang = np.clip(cosang, -1.0, 1.0)
    return float(np.degrees(np.arccos(cosang)))


def _extract_visible_text(decoded_html: str) -> str:
    try:
        from bs4 import BeautifulSoup  # type: ignore

        soup = BeautifulSoup(decoded_html, "html.parser")
        return _normalize_ws(soup.get_text(" ", strip=True))
    except Exception:
        return _normalize_ws(re.sub(r"<[^>]+>", " ", decoded_html))


def _extract_meta_values(decoded_html: str, prop_name: str) -> list[str]:
    return [
        _normalize_ws(value)
        for value in re.findall(
            rf'property=["\']{re.escape(prop_name)}["\']\s+content=["\']([^"\']+)["\']',
            decoded_html,
            flags=re.IGNORECASE,
        )
    ]


def _parse_mp_html() -> dict:
    raw_html = HTML_FILE.read_text(encoding="utf-8", errors="ignore")
    decoded_html = html_lib.unescape(raw_html)
    visible_text = _extract_visible_text(decoded_html)

    title_candidates = _extract_meta_values(decoded_html, "og:title")
    description_candidates = _extract_meta_values(decoded_html, "og:description")

    title_text = next((value for value in title_candidates if "mp-" in value.lower()), "")
    description_text = next(
        (
            value
            for value in description_candidates
            if "two shorter" in value.lower() and "two longer" in value.lower()
        ),
        description_candidates[0] if description_candidates else "",
    )
    text = " ".join(part for part in (title_text, description_text, visible_text) if part)

    material_match = re.search(r"\bmp-\d+\b", text, flags=re.IGNORECASE)
    assert material_match, "Could not extract Materials Project material identifier from HTML"
    material_id = material_match.group(0)

    sg_number = None
    title_tuple_match = re.search(r"\((?:[^,]+,\s*)?([^,]+),\s*([0-9]{1,3})\)", title_text)
    if title_tuple_match:
        sg_number = int(title_tuple_match.group(2))
    if sg_number is None:
        sg_number_match = re.search(r"(?:#|No\.?\s*|number\s*)(\d{1,3})", text, flags=re.IGNORECASE)
        if sg_number_match:
            sg_number = int(sg_number_match.group(1))
    if sg_number is None:
        candidates = [int(value) for value in re.findall(r"\b(1\d{2}|\d{2})\b", text)]
        assert candidates, "Could not extract any integer candidates for space group number from HTML"
        sg_number = 154 if 154 in candidates else candidates[0]

    sg_symbol_raw = None
    if title_tuple_match:
        sg_symbol_raw = title_tuple_match.group(1).split(",")[-1].strip()
    if sg_symbol_raw is None:
        for pattern in (
            r"\bP\s*3(?:_|\s*)?(?:2|\u2082)?\s*(?:2|\u2082)?\s*1\b",
            r"\bP3221\b",
            r"\bP3_221\b",
            r"\bP3\u208221\b",
        ):
            match = re.search(pattern, text, flags=re.IGNORECASE)
            if match:
                sg_symbol_raw = match.group(0)
                break
    assert sg_symbol_raw is not None, "Could not extract space group symbol from HTML"

    all_floats = [float(value) for value in re.findall(r"\b\d+\.\d+\b", text)]
    assert all_floats, "Could not extract any float values from HTML"

    targeted_bonds = re.search(
        r"two\s+shorter\s*\((\d+\.\d+)[^)]*\)\s*and\s*two\s+longer\s*\((\d+\.\d+)[^)]*\)\s*Si-O\s+bond\s+lengths",
        description_text,
        flags=re.IGNORECASE,
    )
    if targeted_bonds:
        bond_candidates = [float(targeted_bonds.group(1)), float(targeted_bonds.group(2))]
    else:
        bond_candidates = []
        si_o_positions = [match.start() for match in re.finditer(r"Si\s*[-\u2013]\s*O", text, flags=re.IGNORECASE)]
        if si_o_positions:
            start = max(0, si_o_positions[0] - 120)
            end = min(len(text), si_o_positions[0] + 240)
            window = text[start:end]
            bond_candidates = [
                float(value)
                for value in re.findall(r"\b\d+\.\d+\b", window)
                if 1.3 <= float(value) <= 2.1
            ]
        if len(bond_candidates) < 2:
            bond_candidates = [value for value in all_floats if 1.3 <= value <= 2.1]

    assert len(bond_candidates) >= 2, "Could not extract two Si-O bond length values from HTML"
    bond_candidates = sorted(bond_candidates)[:2]

    targeted_angle = re.search(
        r"bent\s+(\d+(?:\.\d+)?)\s+degrees\s+geometry",
        description_text,
        flags=re.IGNORECASE,
    )
    if targeted_angle:
        angle_candidates = [float(targeted_angle.group(1))]
    else:
        angle_candidates = []
        angle_label_match = re.search(
            r"(O\s*[-\u2013]\s*Si\s*[-\u2013]\s*O|Si\s*[-\u2013]\s*O\s*[-\u2013]\s*Si)",
            text,
            flags=re.IGNORECASE,
        )
        if angle_label_match:
            start = max(0, angle_label_match.start() - 120)
            end = min(len(text), angle_label_match.end() + 200)
            window = text[start:end]
            angle_candidates = [
                float(value)
                for value in re.findall(r"\b\d+(?:\.\d+)?\b", window)
                if 90 <= float(value) <= 180
            ]
        if not angle_candidates:
            angle_candidates = [value for value in all_floats if 90 <= value <= 180]

    assert angle_candidates, "Could not extract an angle value from HTML"

    return {
        "material_id": material_id,
        "space_group_symbol_raw": sg_symbol_raw,
        "space_group_symbol_norm": _normalize_symbol(sg_symbol_raw),
        "space_group_number": sg_number,
        "si_o_short": float(bond_candidates[0]),
        "si_o_long": float(bond_candidates[1]),
        "angle_deg": float(angle_candidates[0]),
    }


def _compute_structure_metrics(path: Path) -> dict:
    structure = Structure.from_file(str(path))
    sga = SpacegroupAnalyzer(structure, symprec=0.01)
    sg_symbol = sga.get_space_group_symbol()
    sg_number = int(sga.get_space_group_number())

    cnn = CrystalNN()
    short_vals = []
    long_vals = []

    for index, site in enumerate(structure):
        if _element_symbol(site) != "Si":
            continue
        nn_info = cnn.get_nn_info(structure, index)
        o_neighbors = []
        for nn in nn_info:
            nbr_site = nn["site"]
            if _element_symbol(nbr_site) == "O":
                o_neighbors.append(float(site.distance(nbr_site)))

        assert len(o_neighbors) >= 2, f"Si site {index} in {path.name} has fewer than 2 O neighbors"
        o_neighbors = sorted(o_neighbors)
        short_vals.extend(o_neighbors[:2])
        long_vals.extend(o_neighbors[-2:])

    assert short_vals and long_vals, f"No Si-O neighbor distances collected for {path.name}"
    mean_short = float(np.mean(short_vals))
    mean_long = float(np.mean(long_vals))

    angles = []
    for index, site in enumerate(structure):
        if _element_symbol(site) != "O":
            continue
        nn_info = cnn.get_nn_info(structure, index)
        si_neighbors = []
        for nn in nn_info:
            nbr_site = nn["site"]
            if _element_symbol(nbr_site) == "Si":
                si_neighbors.append(nbr_site)

        if len(si_neighbors) == 2:
            a = np.array(si_neighbors[0].coords)
            o = np.array(site.coords)
            b = np.array(si_neighbors[1].coords)
            angles.append(_angle_deg(a, o, b))

    assert angles, f"No O sites with exactly two Si neighbors found for {path.name}"

    return {
        "space_group_symbol_raw": sg_symbol,
        "space_group_symbol_norm": _normalize_symbol(sg_symbol),
        "space_group_number": sg_number,
        "si_o_short": mean_short,
        "si_o_long": mean_long,
        "angle_deg": float(np.mean(angles)),
    }


def _bool_cell_to_python(value):
    if value in (True, False):
        return value
    if isinstance(value, str):
        upper = value.strip().upper()
        if upper == "TRUE":
            return True
        if upper == "FALSE":
            return False
    raise AssertionError(f"Expected TRUE/FALSE boolean-style cell value, got {value!r}")


@pytest.fixture(scope="module")
def expected():
    return {
        "mp": _parse_mp_html(),
        "quartz": _compute_structure_metrics(QUARTZ_CIF),
        "cristobalite": _compute_structure_metrics(CRISTOBALITE_CIF),
    }


@pytest.fixture(scope="module")
def workbook():
    assert OUTPUT_FILE.exists(), f"Output file not found at {OUTPUT_FILE}"
    assert OUTPUT_FILE.stat().st_size > 0, "Output file is empty"
    wb = load_workbook(OUTPUT_FILE)
    yield wb
    wb.close()


@pytest.fixture(scope="module")
def worksheet(workbook):
    assert workbook.sheetnames == [SHEET_NAME], (
        f"Workbook must contain exactly one sheet named {SHEET_NAME!r}, got {workbook.sheetnames}"
    )
    return workbook[SHEET_NAME]


class TestOutputFileExists:
    def test_output_file_exists(self):
        assert OUTPUT_FILE.exists(), f"Output file not found at {OUTPUT_FILE}"

    def test_output_file_not_empty(self):
        assert OUTPUT_FILE.exists(), f"Output file not found at {OUTPUT_FILE}"
        assert OUTPUT_FILE.stat().st_size > 0, "Output file is empty"

    def test_output_file_has_xlsx_extension(self):
        assert OUTPUT_FILE.suffix.lower() == ".xlsx", "Output file must be an .xlsx workbook"


class TestWorkbookStructure:
    def test_exactly_one_sheet_named_verification(self, workbook):
        assert workbook.sheetnames == [SHEET_NAME], (
            f"Workbook must contain exactly one sheet named {SHEET_NAME!r}, got {workbook.sheetnames}"
        )


class TestTableLayout:
    def test_header_row_exact(self, worksheet):
        actual = [worksheet["A1"].value, worksheet["B1"].value, worksheet["C1"].value, worksheet["D1"].value, worksheet["E1"].value]
        assert actual == EXPECTED_HEADERS, f"Header row mismatch. Expected {EXPECTED_HEADERS}, got {actual}"

    def test_metric_labels_exact(self, worksheet):
        actual = [worksheet[f"A{i}"].value for i in range(2, 8)]
        assert actual == EXPECTED_METRICS, f"Metric labels mismatch. Expected {EXPECTED_METRICS}, got {actual}"

    def test_all_main_cells_present(self, worksheet):
        for row in range(1, 8):
            for col in "ABCDE":
                assert worksheet[f"{col}{row}"].value is not None, f"Cell {col}{row} should not be empty"


class TestMPReferenceValues:
    def test_b2_material_id_exact(self, worksheet, expected):
        assert worksheet["B2"].value == expected["mp"]["material_id"]

    def test_b3_space_group_symbol_equivalent(self, worksheet, expected):
        assert _normalize_symbol(worksheet["B3"].value) == expected["mp"]["space_group_symbol_norm"]

    def test_b4_space_group_number_exact(self, worksheet, expected):
        assert int(worksheet["B4"].value) == expected["mp"]["space_group_number"]

    def test_b5_mp_short_distance_numeric(self, worksheet, expected):
        assert float(worksheet["B5"].value) == pytest.approx(expected["mp"]["si_o_short"], rel=0, abs=1e-6)

    def test_b6_mp_long_distance_numeric(self, worksheet, expected):
        assert float(worksheet["B6"].value) == pytest.approx(expected["mp"]["si_o_long"], rel=0, abs=1e-6)

    def test_b7_mp_angle_numeric(self, worksheet, expected):
        assert float(worksheet["B7"].value) == pytest.approx(expected["mp"]["angle_deg"], rel=0, abs=1e-6)


class TestQuartzCIFValues:
    def test_c2_nonempty(self, worksheet):
        assert str(worksheet["C2"].value).strip() != ""

    def test_c3_quartz_space_group_symbol(self, worksheet, expected):
        assert _normalize_symbol(worksheet["C3"].value) == expected["quartz"]["space_group_symbol_norm"]

    def test_c4_quartz_space_group_number(self, worksheet, expected):
        assert int(worksheet["C4"].value) == expected["quartz"]["space_group_number"]

    def test_c5_quartz_short_distance(self, worksheet, expected):
        assert float(worksheet["C5"].value) == pytest.approx(expected["quartz"]["si_o_short"], rel=0, abs=1e-6)

    def test_c6_quartz_long_distance(self, worksheet, expected):
        assert float(worksheet["C6"].value) == pytest.approx(expected["quartz"]["si_o_long"], rel=0, abs=1e-6)

    def test_c7_quartz_angle(self, worksheet, expected):
        assert float(worksheet["C7"].value) == pytest.approx(expected["quartz"]["angle_deg"], rel=0, abs=1e-6)


class TestCristobaliteCIFValues:
    def test_d2_nonempty(self, worksheet):
        assert str(worksheet["D2"].value).strip() != ""

    def test_d3_cristobalite_space_group_symbol(self, worksheet, expected):
        assert _normalize_symbol(worksheet["D3"].value) == expected["cristobalite"]["space_group_symbol_norm"]

    def test_d4_cristobalite_space_group_number(self, worksheet, expected):
        assert int(worksheet["D4"].value) == expected["cristobalite"]["space_group_number"]

    def test_d5_cristobalite_short_distance(self, worksheet, expected):
        assert float(worksheet["D5"].value) == pytest.approx(expected["cristobalite"]["si_o_short"], rel=0, abs=1e-6)

    def test_d6_cristobalite_long_distance(self, worksheet, expected):
        assert float(worksheet["D6"].value) == pytest.approx(expected["cristobalite"]["si_o_long"], rel=0, abs=1e-6)

    def test_d7_cristobalite_angle(self, worksheet, expected):
        assert float(worksheet["D7"].value) == pytest.approx(expected["cristobalite"]["angle_deg"], rel=0, abs=1e-6)


class TestMatchColumn:
    def test_e2_material_id_match_logic(self, worksheet):
        expected_match = str(worksheet["B2"].value) == str(worksheet["C2"].value)
        assert _bool_cell_to_python(worksheet["E2"].value) == expected_match

    def test_e3_space_group_symbol_match_logic(self, worksheet, expected):
        expected_match = expected["mp"]["space_group_symbol_norm"] == expected["quartz"]["space_group_symbol_norm"]
        assert _bool_cell_to_python(worksheet["E3"].value) == expected_match

    def test_e4_space_group_number_match_logic(self, worksheet, expected):
        expected_match = expected["mp"]["space_group_number"] == expected["quartz"]["space_group_number"]
        assert _bool_cell_to_python(worksheet["E4"].value) == expected_match

    def test_e5_short_distance_tolerance_logic(self, worksheet, expected):
        diff = abs(expected["mp"]["si_o_short"] - expected["quartz"]["si_o_short"])
        assert _bool_cell_to_python(worksheet["E5"].value) == (diff <= 0.02)

    def test_e6_long_distance_tolerance_logic(self, worksheet, expected):
        diff = abs(expected["mp"]["si_o_long"] - expected["quartz"]["si_o_long"])
        assert _bool_cell_to_python(worksheet["E6"].value) == (diff <= 0.02)

    def test_e7_angle_tolerance_logic(self, worksheet, expected):
        diff = abs(expected["mp"]["angle_deg"] - expected["quartz"]["angle_deg"])
        assert _bool_cell_to_python(worksheet["E7"].value) == (diff <= 2.0)


class TestFormatting:
    def test_header_row_bold_all_cells(self, worksheet):
        for cell in worksheet[1]:
            assert cell.font.bold is True, f"Header cell {cell.coordinate} must be bold"

    def test_mp_column_blue_font(self, worksheet):
        for row in range(2, 8):
            assert _font_is_blue(worksheet[f"B{row}"]), f"Cell B{row} must use blue font ({BLUE_RGB})"

    def test_cif_columns_black_font(self, worksheet):
        for col in ("C", "D"):
            for row in range(2, 8):
                assert _font_is_black_or_default(worksheet[f"{col}{row}"]), (
                    f"Cell {col}{row} must use black/default font ({BLACK_RGB})"
                )

    def test_distance_number_formats(self, worksheet):
        for col in ("B", "C", "D"):
            for row in (5, 6):
                assert worksheet[f"{col}{row}"].number_format == "0.000"

    def test_angle_number_formats(self, worksheet):
        for col in ("B", "C", "D"):
            assert worksheet[f"{col}7"].number_format == "0.0"


class TestSanityChecks:
    def test_quartz_and_cristobalite_space_groups_differ(self, expected):
        assert expected["quartz"]["space_group_number"] != expected["cristobalite"]["space_group_number"]

    def test_short_leq_long_all_columns(self, worksheet):
        for col in ("B", "C", "D"):
            assert float(worksheet[f"{col}5"].value) <= float(worksheet[f"{col}6"].value) + 1e-9
