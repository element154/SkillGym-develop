"""Auto-generated expectation tests for task verification.

These tests verify that the task execution produces correct outputs
based on the instruction requirements for the median account report task.

Expected calculations:
- Target account: 146832 (closest to median net revenue)
- Customer name: Kiehn-Spinka
- Net Revenue 2014: 99608.77
- Median Net Revenue (all accounts): 100271.535
- Abs Diff from Median: ~662.765
- Weighted Avg Unit Price (positive qty only): ~56.723
- Catalog Median Product Price: 19.5
- AboveCatalogMedian: True (because 56.723 > 19.5)
"""

import os
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


class TestOutputFileExists:
    """Tests for output file existence and basic structure."""

    OUTPUT_PATH = "/root/median_account_report.xlsx"

    def test_output_file_exists(self):
        """Verify output file was created at the specified path."""
        assert os.path.exists(self.OUTPUT_PATH), (
            f"Output file not found at {self.OUTPUT_PATH}"
        )

    def test_output_file_is_valid_excel(self):
        """Verify output is a valid Excel file that can be loaded."""
        wb = load_workbook(self.OUTPUT_PATH)
        assert wb is not None, "Failed to load output file as Excel workbook"
        wb.close()

    def test_output_directory_exists(self):
        """Verify output directory exists."""
        output_dir = "/root"
        assert os.path.isdir(output_dir), f"Output directory not found: {output_dir}"


class TestWorkbookStructure:
    """Tests for workbook sheet structure and template preservation."""

    OUTPUT_PATH = "/root/median_account_report.xlsx"
    TEMPLATE_SHEETS = [
        'Max-Min', 'IF-IFS', 'Len', 'LeftRight', 'DateToText', 'TRIM',
        'Substitute', 'SUM-SumIF', 'Count-CountIF', 'Concatenate', 'Days-NetworkDays'
    ]

    def test_median_account_sheet_exists(self):
        """Verify MedianAccount sheet was created."""
        wb = load_workbook(self.OUTPUT_PATH)
        assert 'MedianAccount' in wb.sheetnames, (
            f"MedianAccount sheet not found. Available sheets: {wb.sheetnames}"
        )
        wb.close()

    def test_template_sheets_preserved(self):
        """Verify all original template sheets are preserved."""
        wb = load_workbook(self.OUTPUT_PATH)
        for sheet_name in self.TEMPLATE_SHEETS:
            assert sheet_name in wb.sheetnames, (
                f"Template sheet '{sheet_name}' is missing from output workbook"
            )
        wb.close()

    def test_total_sheet_count(self):
        """Verify correct total number of sheets (11 template + 1 MedianAccount)."""
        wb = load_workbook(self.OUTPUT_PATH)
        expected_count = len(self.TEMPLATE_SHEETS) + 1  # 12 total
        assert len(wb.sheetnames) == expected_count, (
            f"Expected {expected_count} sheets, found {len(wb.sheetnames)}: {wb.sheetnames}"
        )
        wb.close()


class TestMedianAccountSheetHeaders:
    """Tests for MedianAccount sheet header structure."""

    OUTPUT_PATH = "/root/median_account_report.xlsx"
    REQUIRED_HEADERS = [
        'AccountNumber',
        'CustomerName',
        'NetRevenue2014',
        'MedianNetRevenueAllAccounts',
        'AbsDiffFromMedian',
        'WeightedAvgUnitPrice_PosQtyOnly',
        'CatalogMedianProductPrice',
        'AboveCatalogMedian'
    ]

    def test_required_headers_present(self):
        """Verify all required column headers are present in MedianAccount sheet."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb['MedianAccount']

        # Get all values from the first row (headers)
        headers = []
        for cell in ws[1]:
            if cell.value is not None:
                headers.append(str(cell.value))

        for required_header in self.REQUIRED_HEADERS:
            assert required_header in headers, (
                f"Required header '{required_header}' not found. Found headers: {headers}"
            )
        wb.close()

    def test_header_count_matches(self):
        """Verify the number of headers matches expected count."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb['MedianAccount']

        headers = [cell.value for cell in ws[1] if cell.value is not None]
        assert len(headers) == len(self.REQUIRED_HEADERS), (
            f"Expected {len(self.REQUIRED_HEADERS)} headers, found {len(headers)}: {headers}"
        )
        wb.close()


class TestMedianAccountValues:
    """Tests for computed values in MedianAccount sheet."""

    OUTPUT_PATH = "/root/median_account_report.xlsx"

    # Expected values (computed from task requirements)
    EXPECTED_ACCOUNT_NUMBER = 146832
    EXPECTED_CUSTOMER_NAME = "Kiehn-Spinka"
    EXPECTED_NET_REVENUE = 99608.77
    EXPECTED_MEDIAN_NET_REVENUE = 100271.535
    EXPECTED_ABS_DIFF = 662.765  # approximately
    EXPECTED_WEIGHTED_AVG_PRICE = 56.723  # approximately
    EXPECTED_CATALOG_MEDIAN = 19.5
    EXPECTED_ABOVE_CATALOG_MEDIAN = True

    def _get_sheet_data(self):
        """Helper to load MedianAccount data as a dictionary."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb['MedianAccount']

        # Get headers from row 1
        headers = [cell.value for cell in ws[1] if cell.value is not None]

        # Get values from row 2 (data row)
        values = [cell.value for cell in ws[2]][:len(headers)]

        data = dict(zip(headers, values))
        wb.close()
        return data

    def test_account_number_value(self):
        """Verify AccountNumber is the expected value (146832)."""
        data = self._get_sheet_data()
        account_num = data.get('AccountNumber')
        assert account_num is not None, "AccountNumber value is missing"
        assert int(account_num) == self.EXPECTED_ACCOUNT_NUMBER, (
            f"Expected AccountNumber {self.EXPECTED_ACCOUNT_NUMBER}, got {account_num}"
        )

    def test_customer_name_value(self):
        """Verify CustomerName is the expected value."""
        data = self._get_sheet_data()
        customer_name = data.get('CustomerName')
        assert customer_name is not None, "CustomerName value is missing"
        assert str(customer_name).strip() == self.EXPECTED_CUSTOMER_NAME, (
            f"Expected CustomerName '{self.EXPECTED_CUSTOMER_NAME}', got '{customer_name}'"
        )

    def test_net_revenue_2014_value(self):
        """Verify NetRevenue2014 is close to expected value."""
        data = self._get_sheet_data()
        net_revenue = data.get('NetRevenue2014')
        assert net_revenue is not None, "NetRevenue2014 value is missing"
        assert abs(float(net_revenue) - self.EXPECTED_NET_REVENUE) < 0.01, (
            f"Expected NetRevenue2014 ~{self.EXPECTED_NET_REVENUE}, got {net_revenue}"
        )

    def test_median_net_revenue_value(self):
        """Verify MedianNetRevenueAllAccounts is close to expected value."""
        data = self._get_sheet_data()
        median_rev = data.get('MedianNetRevenueAllAccounts')
        assert median_rev is not None, "MedianNetRevenueAllAccounts value is missing"
        assert abs(float(median_rev) - self.EXPECTED_MEDIAN_NET_REVENUE) < 0.01, (
            f"Expected MedianNetRevenueAllAccounts ~{self.EXPECTED_MEDIAN_NET_REVENUE}, got {median_rev}"
        )

    def test_abs_diff_from_median_value(self):
        """Verify AbsDiffFromMedian is close to expected value."""
        data = self._get_sheet_data()
        abs_diff = data.get('AbsDiffFromMedian')
        assert abs_diff is not None, "AbsDiffFromMedian value is missing"
        assert abs(float(abs_diff) - self.EXPECTED_ABS_DIFF) < 0.01, (
            f"Expected AbsDiffFromMedian ~{self.EXPECTED_ABS_DIFF}, got {abs_diff}"
        )

    def test_weighted_avg_unit_price_value(self):
        """Verify WeightedAvgUnitPrice_PosQtyOnly is close to expected value."""
        data = self._get_sheet_data()
        weighted_avg = data.get('WeightedAvgUnitPrice_PosQtyOnly')
        assert weighted_avg is not None, "WeightedAvgUnitPrice_PosQtyOnly value is missing"
        assert abs(float(weighted_avg) - self.EXPECTED_WEIGHTED_AVG_PRICE) < 0.01, (
            f"Expected WeightedAvgUnitPrice_PosQtyOnly ~{self.EXPECTED_WEIGHTED_AVG_PRICE}, got {weighted_avg}"
        )

    def test_catalog_median_product_price_value(self):
        """Verify CatalogMedianProductPrice is the expected value."""
        data = self._get_sheet_data()
        catalog_median = data.get('CatalogMedianProductPrice')
        assert catalog_median is not None, "CatalogMedianProductPrice value is missing"
        assert abs(float(catalog_median) - self.EXPECTED_CATALOG_MEDIAN) < 0.01, (
            f"Expected CatalogMedianProductPrice ~{self.EXPECTED_CATALOG_MEDIAN}, got {catalog_median}"
        )

    def test_above_catalog_median_value(self):
        """Verify AboveCatalogMedian is TRUE (since weighted avg > catalog median)."""
        data = self._get_sheet_data()
        above_flag = data.get('AboveCatalogMedian')
        assert above_flag is not None, "AboveCatalogMedian value is missing"

        # Handle various representations of True
        if isinstance(above_flag, bool):
            assert above_flag is True, f"Expected AboveCatalogMedian to be True, got {above_flag}"
        elif isinstance(above_flag, str):
            assert above_flag.upper() in ['TRUE', 'YES', '1'], (
                f"Expected AboveCatalogMedian to be TRUE, got '{above_flag}'"
            )
        else:
            assert bool(above_flag), f"Expected AboveCatalogMedian to be truthy, got {above_flag}"


class TestValueConsistency:
    """Tests for internal consistency of computed values."""

    OUTPUT_PATH = "/root/median_account_report.xlsx"

    def _get_sheet_data(self):
        """Helper to load MedianAccount data as a dictionary."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb['MedianAccount']
        headers = [cell.value for cell in ws[1] if cell.value is not None]
        values = [cell.value for cell in ws[2]][:len(headers)]
        data = dict(zip(headers, values))
        wb.close()
        return data

    def test_abs_diff_calculation_correct(self):
        """Verify AbsDiffFromMedian = |NetRevenue2014 - MedianNetRevenueAllAccounts|."""
        data = self._get_sheet_data()
        net_rev = float(data['NetRevenue2014'])
        median_rev = float(data['MedianNetRevenueAllAccounts'])
        abs_diff = float(data['AbsDiffFromMedian'])

        expected_abs_diff = abs(net_rev - median_rev)
        assert abs(abs_diff - expected_abs_diff) < 0.01, (
            f"AbsDiffFromMedian ({abs_diff}) should equal |{net_rev} - {median_rev}| = {expected_abs_diff}"
        )

    def test_above_catalog_median_calculation_correct(self):
        """Verify AboveCatalogMedian flag matches weighted avg > catalog median."""
        data = self._get_sheet_data()
        weighted_avg = float(data['WeightedAvgUnitPrice_PosQtyOnly'])
        catalog_median = float(data['CatalogMedianProductPrice'])
        above_flag = data['AboveCatalogMedian']

        expected_flag = weighted_avg > catalog_median

        # Normalize the flag to boolean
        if isinstance(above_flag, bool):
            actual_flag = above_flag
        elif isinstance(above_flag, str):
            actual_flag = above_flag.upper() in ['TRUE', 'YES', '1']
        else:
            actual_flag = bool(above_flag)

        assert actual_flag == expected_flag, (
            f"AboveCatalogMedian should be {expected_flag} "
            f"(weighted avg {weighted_avg} {'>' if expected_flag else '<='} catalog median {catalog_median})"
        )


class TestHeaderFormatting:
    """Tests for header formatting based on sample-1-sheet.xlsx boolean value."""

    OUTPUT_PATH = "/root/median_account_report.xlsx"

    def test_header_row_is_bold(self):
        """Verify header row has bold font (since boolean=True for number=1)."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb['MedianAccount']

        # Check if any header cell is bold (at least one should be)
        has_bold = False
        for cell in ws[1]:
            if cell.value is not None and cell.font and cell.font.bold:
                has_bold = True
                break

        assert has_bold, (
            "Header row should have bold font because boolean=True for number=1 "
            "in sample-1-sheet.xlsx"
        )
        wb.close()

    def test_header_row_has_yellow_fill(self):
        """Verify header row has yellow fill (since boolean=True for number=1)."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb['MedianAccount']

        # Check if any header cell has a yellow fill
        has_yellow_fill = False
        for cell in ws[1]:
            if cell.value is not None and cell.fill:
                # Check for yellow fill (various shades)
                fill_color = cell.fill.start_color
                if fill_color and fill_color.rgb:
                    # Yellow colors typically have high R, high G, low B
                    # Common yellow: FFFF00, FFFF99, etc.
                    rgb = str(fill_color.rgb)
                    if rgb.upper() in ['FFFFFF00', 'FFFF00', 'FFFFFF99', 'FFFFCC00', 'FFFFE599'] or \
                       (len(rgb) >= 6 and rgb[-6:-4].upper() == 'FF' and rgb[-4:-2].upper() == 'FF'):
                        has_yellow_fill = True
                        break
                    # Also check for any non-white fill that could be yellow
                    if fill_color.type == 'rgb' and fill_color.rgb not in [None, '00000000', 'FFFFFFFF']:
                        # Check if it's a yellowish color
                        try:
                            if len(rgb) >= 6:
                                r = int(rgb[-6:-4], 16) if len(rgb) >= 6 else 0
                                g = int(rgb[-4:-2], 16) if len(rgb) >= 4 else 0
                                b = int(rgb[-2:], 16) if len(rgb) >= 2 else 0
                                # Yellow has high R, high G, low B
                                if r > 200 and g > 200 and b < 150:
                                    has_yellow_fill = True
                                    break
                        except (ValueError, IndexError):
                            pass

        assert has_yellow_fill, (
            "Header row should have yellow fill because boolean=True for number=1 "
            "in sample-1-sheet.xlsx"
        )
        wb.close()


class TestDataRowStructure:
    """Tests for data row structure in MedianAccount sheet."""

    OUTPUT_PATH = "/root/median_account_report.xlsx"

    def test_single_data_row_only(self):
        """Verify there is exactly one data row (plus header)."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb['MedianAccount']

        # Count non-empty rows
        non_empty_rows = 0
        for row in ws.iter_rows(max_row=10):  # Check first 10 rows
            if any(cell.value is not None for cell in row):
                non_empty_rows += 1

        # Should have exactly 2 rows: header + 1 data row
        assert non_empty_rows == 2, (
            f"Expected exactly 2 rows (1 header + 1 data), found {non_empty_rows} non-empty rows"
        )
        wb.close()

    def test_all_values_populated(self):
        """Verify all required fields have values (no None/empty)."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb['MedianAccount']

        headers = [cell.value for cell in ws[1] if cell.value is not None]
        values = [cell.value for cell in ws[2]][:len(headers)]

        for i, (header, value) in enumerate(zip(headers, values)):
            assert value is not None, f"Value for '{header}' is missing (None)"
            if isinstance(value, str):
                assert value.strip() != '', f"Value for '{header}' is empty string"

        wb.close()


class TestNumericTypes:
    """Tests for correct numeric types in computed values."""

    OUTPUT_PATH = "/root/median_account_report.xlsx"

    def _get_sheet_data(self):
        """Helper to load MedianAccount data as a dictionary."""
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb['MedianAccount']
        headers = [cell.value for cell in ws[1] if cell.value is not None]
        values = [cell.value for cell in ws[2]][:len(headers)]
        data = dict(zip(headers, values))
        wb.close()
        return data

    def test_account_number_is_numeric(self):
        """Verify AccountNumber can be converted to integer."""
        data = self._get_sheet_data()
        account_num = data.get('AccountNumber')
        try:
            int(account_num)
        except (TypeError, ValueError) as e:
            pytest.fail(f"AccountNumber '{account_num}' cannot be converted to integer: {e}")

    def test_revenue_values_are_numeric(self):
        """Verify revenue-related values are numeric."""
        data = self._get_sheet_data()
        numeric_fields = [
            'NetRevenue2014',
            'MedianNetRevenueAllAccounts',
            'AbsDiffFromMedian',
            'WeightedAvgUnitPrice_PosQtyOnly',
            'CatalogMedianProductPrice'
        ]

        for field in numeric_fields:
            value = data.get(field)
            try:
                float(value)
            except (TypeError, ValueError) as e:
                pytest.fail(f"{field} value '{value}' cannot be converted to float: {e}")


class TestSourceDataValidation:
    """Tests to validate that output data is consistent with source files."""

    OUTPUT_PATH = "/root/median_account_report.xlsx"
    INPUT_DIR = "/root"

    def test_account_number_exists_in_sales_data(self):
        """Verify the reported account number exists in the sales data."""
        # Get account from output
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb['MedianAccount']
        headers = [cell.value for cell in ws[1] if cell.value is not None]
        values = [cell.value for cell in ws[2]][:len(headers)]
        data = dict(zip(headers, values))
        wb.close()

        account_num = int(data['AccountNumber'])

        # Check against source
        sales_df = pd.read_excel(f"{self.INPUT_DIR}/sample-salesv3.xlsx")
        assert account_num in sales_df['account number'].values, (
            f"AccountNumber {account_num} not found in sample-salesv3.xlsx"
        )

    def test_customer_name_exists_for_account(self):
        """Verify the reported customer name exists for the account in sales data."""
        # Get data from output
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb['MedianAccount']
        headers = [cell.value for cell in ws[1] if cell.value is not None]
        values = [cell.value for cell in ws[2]][:len(headers)]
        data = dict(zip(headers, values))
        wb.close()

        account_num = int(data['AccountNumber'])
        customer_name = str(data['CustomerName']).strip()

        # Check against source
        sales_df = pd.read_excel(f"{self.INPUT_DIR}/sample-salesv3.xlsx")
        account_names = sales_df[sales_df['account number'] == account_num]['name'].unique()

        assert customer_name in account_names, (
            f"CustomerName '{customer_name}' not found for account {account_num}. "
            f"Available names: {list(account_names)}"
        )

    def test_catalog_median_matches_products_file(self):
        """Verify CatalogMedianProductPrice matches median from Products.xlsx."""
        # Get data from output
        wb = load_workbook(self.OUTPUT_PATH)
        ws = wb['MedianAccount']
        headers = [cell.value for cell in ws[1] if cell.value is not None]
        values = [cell.value for cell in ws[2]][:len(headers)]
        data = dict(zip(headers, values))
        wb.close()

        reported_median = float(data['CatalogMedianProductPrice'])

        # Calculate from source
        products_df = pd.read_excel(f"{self.INPUT_DIR}/Products.xlsx")
        actual_median = products_df['Price'].median()

        assert abs(reported_median - actual_median) < 0.01, (
            f"CatalogMedianProductPrice ({reported_median}) doesn't match "
            f"actual median from Products.xlsx ({actual_median})"
        )


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
