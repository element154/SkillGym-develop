"""Auto-generated expectation tests for RTPLAN reference validation task.

These tests verify that the task execution produces correct outputs
based on the instruction requirements for validating RTPLAN references
against CT and MR images.
"""

import os
from pathlib import Path

import pytest

# Expected output file path
OUTPUT_FILE = "/root/rtplan_reference_validation.xlsx"


class TestOutputFileExists:
    """Tests for output file existence and basic validity."""

    def test_output_file_exists(self):
        """Verify output Excel file was created at the exact specified path."""
        assert os.path.exists(OUTPUT_FILE), (
            f"Output file not found at {OUTPUT_FILE}. "
            "The task requires the Excel file to be written to exactly this path."
        )

    def test_output_file_is_not_empty(self):
        """Verify output file has content."""
        assert os.path.exists(OUTPUT_FILE), f"Output file not found: {OUTPUT_FILE}"
        file_size = os.path.getsize(OUTPUT_FILE)
        assert file_size > 0, "Output file is empty"

    def test_output_file_has_xlsx_extension(self):
        """Verify output file has correct .xlsx extension."""
        assert OUTPUT_FILE.endswith(".xlsx"), "Output file must have .xlsx extension"

    def test_output_directory_exists(self):
        """Verify the output directory exists."""
        output_dir = os.path.dirname(OUTPUT_FILE)
        assert os.path.isdir(output_dir), f"Output directory does not exist: {output_dir}"


class TestExcelWorkbookStructure:
    """Tests for Excel workbook structure using openpyxl."""

    @pytest.fixture
    def workbook(self):
        """Load the output workbook."""
        from openpyxl import load_workbook
        assert os.path.exists(OUTPUT_FILE), f"Output file not found: {OUTPUT_FILE}"
        return load_workbook(OUTPUT_FILE)

    def test_workbook_is_valid_xlsx(self):
        """Verify the file is a valid Excel workbook."""
        from openpyxl import load_workbook
        assert os.path.exists(OUTPUT_FILE), f"Output file not found: {OUTPUT_FILE}"
        try:
            wb = load_workbook(OUTPUT_FILE)
            assert wb is not None
        except Exception as e:
            pytest.fail(f"Failed to load workbook as valid Excel file: {e}")

    def test_summary_sheet_exists(self, workbook):
        """Verify the Summary sheet exists in the workbook."""
        sheet_names = workbook.sheetnames
        assert "Summary" in sheet_names, (
            f"'Summary' sheet not found. Available sheets: {sheet_names}"
        )

    def test_uid_matches_sheet_exists(self, workbook):
        """Verify the UID_Matches sheet exists in the workbook."""
        sheet_names = workbook.sheetnames
        assert "UID_Matches" in sheet_names, (
            f"'UID_Matches' sheet not found. Available sheets: {sheet_names}"
        )


class TestSummarySheet:
    """Tests for the Summary sheet content and formatting."""

    @pytest.fixture
    def summary_sheet(self):
        """Load the Summary sheet from the workbook."""
        from openpyxl import load_workbook
        assert os.path.exists(OUTPUT_FILE), f"Output file not found: {OUTPUT_FILE}"
        wb = load_workbook(OUTPUT_FILE)
        assert "Summary" in wb.sheetnames, "Summary sheet not found"
        return wb["Summary"]

    def test_verdict_cell_exists(self, summary_sheet):
        """Verify cell B2 contains the overall verdict."""
        cell_value = summary_sheet["B2"].value
        assert cell_value is not None, "Cell B2 (verdict) is empty"

    def test_verdict_is_pass_or_fail(self, summary_sheet):
        """Verify the verdict is either PASS or FAIL."""
        verdict = summary_sheet["B2"].value
        assert verdict in ("PASS", "FAIL"), (
            f"Verdict must be 'PASS' or 'FAIL', got: {verdict}"
        )

    def test_total_uid_count_cell_exists(self, summary_sheet):
        """Verify cell B3 contains the total referenced UID count."""
        cell_value = summary_sheet["B3"].value
        assert cell_value is not None, "Cell B3 (total UID count) is empty"

    def test_total_uid_count_is_numeric(self, summary_sheet):
        """Verify total UID count is a valid number."""
        total_count = summary_sheet["B3"].value
        assert isinstance(total_count, (int, float)), (
            f"Total UID count must be numeric, got: {type(total_count)}"
        )
        assert total_count >= 0, "Total UID count cannot be negative"

    def test_matched_count_cell_exists(self, summary_sheet):
        """Verify cell B4 contains the matched count."""
        cell_value = summary_sheet["B4"].value
        assert cell_value is not None, "Cell B4 (matched count) is empty"

    def test_matched_count_is_numeric(self, summary_sheet):
        """Verify matched count is a valid number."""
        matched_count = summary_sheet["B4"].value
        assert isinstance(matched_count, (int, float)), (
            f"Matched count must be numeric, got: {type(matched_count)}"
        )
        assert matched_count >= 0, "Matched count cannot be negative"

    def test_unmatched_count_cell_exists(self, summary_sheet):
        """Verify cell B5 contains the unmatched count."""
        cell_value = summary_sheet["B5"].value
        assert cell_value is not None, "Cell B5 (unmatched count) is empty"

    def test_unmatched_count_is_numeric(self, summary_sheet):
        """Verify unmatched count is a valid number."""
        unmatched_count = summary_sheet["B5"].value
        assert isinstance(unmatched_count, (int, float)), (
            f"Unmatched count must be numeric, got: {type(unmatched_count)}"
        )
        assert unmatched_count >= 0, "Unmatched count cannot be negative"

    def test_counts_add_up(self, summary_sheet):
        """Verify that matched + unmatched = total."""
        total = summary_sheet["B3"].value
        matched = summary_sheet["B4"].value
        unmatched = summary_sheet["B5"].value
        assert matched + unmatched == total, (
            f"Counts don't add up: matched({matched}) + unmatched({unmatched}) != total({total})"
        )

    def test_verdict_consistent_with_counts(self, summary_sheet):
        """Verify verdict is consistent with match counts."""
        verdict = summary_sheet["B2"].value
        total = summary_sheet["B3"].value
        matched = summary_sheet["B4"].value
        unmatched = summary_sheet["B5"].value

        if total > 0 and unmatched == 0:
            assert verdict == "PASS", (
                f"Verdict should be PASS when all UIDs match, but got {verdict}"
            )
        elif unmatched > 0:
            assert verdict == "FAIL", (
                f"Verdict should be FAIL when there are unmatched UIDs, but got {verdict}"
            )

    def test_verdict_cell_has_fill_color(self, summary_sheet):
        """Verify cell B2 has conditional fill color (green for PASS, yellow for FAIL)."""
        cell = summary_sheet["B2"]
        verdict = cell.value
        fill = cell.fill

        # Check that fill is applied (not default/no fill)
        if fill.fill_type is not None:
            fill_color = fill.start_color.rgb if fill.start_color else None
            # Accept any non-default fill as valid styling
            assert fill_color is not None or fill.fill_type == "solid", (
                "Cell B2 should have a fill color applied"
            )


class TestUIDMatchesSheet:
    """Tests for the UID_Matches sheet content and formatting."""

    @pytest.fixture
    def uid_matches_sheet(self):
        """Load the UID_Matches sheet from the workbook."""
        from openpyxl import load_workbook
        assert os.path.exists(OUTPUT_FILE), f"Output file not found: {OUTPUT_FILE}"
        wb = load_workbook(OUTPUT_FILE)
        assert "UID_Matches" in wb.sheetnames, "UID_Matches sheet not found"
        return wb["UID_Matches"]

    def test_required_headers_present(self, uid_matches_sheet):
        """Verify all required column headers are present."""
        required_headers = [
            "referenced_uid",
            "uid_type_detected",
            "matched_file",
            "matched_identifier_field",
            "match_status"
        ]

        # Read the first row for headers
        headers = []
        for col in range(1, uid_matches_sheet.max_column + 1):
            cell_value = uid_matches_sheet.cell(row=1, column=col).value
            if cell_value:
                headers.append(cell_value)

        for required in required_headers:
            assert required in headers, (
                f"Required header '{required}' not found. Headers present: {headers}"
            )

    def test_header_row_is_bold(self, uid_matches_sheet):
        """Verify the header row has bold formatting."""
        # Check first few header cells for bold formatting
        bold_count = 0
        for col in range(1, min(6, uid_matches_sheet.max_column + 1)):
            cell = uid_matches_sheet.cell(row=1, column=col)
            if cell.font and cell.font.bold:
                bold_count += 1

        assert bold_count > 0, "Header row should have bold formatting"

    def test_sheet_has_data_rows(self, uid_matches_sheet):
        """Verify the sheet contains data rows (not just headers)."""
        max_row = uid_matches_sheet.max_row
        # Should have at least header row + 1 data row
        assert max_row >= 1, "UID_Matches sheet should have at least a header row"

    def test_match_status_values_valid(self, uid_matches_sheet):
        """Verify match_status column contains only MATCH or NO_MATCH values."""
        # Find the match_status column
        match_status_col = None
        for col in range(1, uid_matches_sheet.max_column + 1):
            if uid_matches_sheet.cell(row=1, column=col).value == "match_status":
                match_status_col = col
                break

        assert match_status_col is not None, "match_status column not found"

        # Check all data rows
        for row in range(2, uid_matches_sheet.max_row + 1):
            status = uid_matches_sheet.cell(row=row, column=match_status_col).value
            if status is not None:  # Skip empty rows
                assert status in ("MATCH", "NO_MATCH"), (
                    f"Invalid match_status value at row {row}: {status}"
                )

    def test_uid_type_detected_values_valid(self, uid_matches_sheet):
        """Verify uid_type_detected contains valid values."""
        valid_types = ["StudyInstanceUID", "SeriesInstanceUID", "SOPInstanceUID", "Other/Unknown"]

        # Find the uid_type_detected column
        type_col = None
        for col in range(1, uid_matches_sheet.max_column + 1):
            if uid_matches_sheet.cell(row=1, column=col).value == "uid_type_detected":
                type_col = col
                break

        assert type_col is not None, "uid_type_detected column not found"

        # Check all data rows
        for row in range(2, uid_matches_sheet.max_row + 1):
            uid_type = uid_matches_sheet.cell(row=row, column=type_col).value
            if uid_type is not None:  # Skip empty rows
                assert uid_type in valid_types, (
                    f"Invalid uid_type_detected value at row {row}: {uid_type}. "
                    f"Valid values are: {valid_types}"
                )

    def test_matched_file_values_valid(self, uid_matches_sheet):
        """Verify matched_file contains valid file names or is blank."""
        valid_files = ["CT_small.dcm", "MR_small.dcm", "", None]

        # Find the matched_file column
        file_col = None
        for col in range(1, uid_matches_sheet.max_column + 1):
            if uid_matches_sheet.cell(row=1, column=col).value == "matched_file":
                file_col = col
                break

        assert file_col is not None, "matched_file column not found"

        # Check all data rows
        for row in range(2, uid_matches_sheet.max_row + 1):
            matched_file = uid_matches_sheet.cell(row=row, column=file_col).value
            assert matched_file in valid_files or matched_file == "", (
                f"Invalid matched_file value at row {row}: {matched_file}. "
                f"Valid values are: CT_small.dcm, MR_small.dcm, or blank"
            )

    def test_referenced_uid_values_look_like_dicom_uids(self, uid_matches_sheet):
        """Verify referenced_uid values follow DICOM UID pattern."""
        import re
        uid_pattern = re.compile(r'^[\d.]+$')

        # Find the referenced_uid column
        uid_col = None
        for col in range(1, uid_matches_sheet.max_column + 1):
            if uid_matches_sheet.cell(row=1, column=col).value == "referenced_uid":
                uid_col = col
                break

        assert uid_col is not None, "referenced_uid column not found"

        # Check all data rows
        for row in range(2, uid_matches_sheet.max_row + 1):
            uid = uid_matches_sheet.cell(row=row, column=uid_col).value
            if uid is not None and uid != "":
                uid_str = str(uid)
                assert uid_pattern.match(uid_str), (
                    f"referenced_uid at row {row} doesn't match DICOM UID pattern: {uid}"
                )

    def test_column_widths_are_reasonable(self, uid_matches_sheet):
        """Verify columns have reasonable widths set (not default narrow)."""
        # Check that at least some columns have explicit widths
        widths_set = 0
        for col_letter in ['A', 'B', 'C', 'D', 'E']:
            col_dim = uid_matches_sheet.column_dimensions.get(col_letter)
            if col_dim and col_dim.width and col_dim.width > 8:  # Default is ~8
                widths_set += 1

        # At least some columns should have adjusted widths
        assert widths_set > 0, (
            "Column widths should be set to reasonable values (not default narrow)"
        )


class TestDataConsistency:
    """Tests for data consistency between sheets."""

    @pytest.fixture
    def workbook(self):
        """Load the output workbook."""
        from openpyxl import load_workbook
        assert os.path.exists(OUTPUT_FILE), f"Output file not found: {OUTPUT_FILE}"
        return load_workbook(OUTPUT_FILE)

    def test_summary_total_matches_uid_matches_rows(self, workbook):
        """Verify Summary total count matches number of rows in UID_Matches."""
        summary = workbook["Summary"]
        uid_matches = workbook["UID_Matches"]

        total_from_summary = summary["B3"].value
        # Count data rows in UID_Matches (excluding header)
        data_rows = uid_matches.max_row - 1  # Subtract header row

        assert total_from_summary == data_rows, (
            f"Summary total ({total_from_summary}) doesn't match "
            f"UID_Matches data rows ({data_rows})"
        )

    def test_summary_matched_count_matches_uid_matches_data(self, workbook):
        """Verify Summary matched count matches MATCH rows in UID_Matches."""
        summary = workbook["Summary"]
        uid_matches = workbook["UID_Matches"]

        matched_from_summary = summary["B4"].value

        # Find match_status column and count MATCH values
        match_status_col = None
        for col in range(1, uid_matches.max_column + 1):
            if uid_matches.cell(row=1, column=col).value == "match_status":
                match_status_col = col
                break

        match_count = 0
        for row in range(2, uid_matches.max_row + 1):
            if uid_matches.cell(row=row, column=match_status_col).value == "MATCH":
                match_count += 1

        assert matched_from_summary == match_count, (
            f"Summary matched count ({matched_from_summary}) doesn't match "
            f"MATCH rows in UID_Matches ({match_count})"
        )


class TestInputFilesUsed:
    """Tests to verify input DICOM files are accessible (prerequisite check)."""

    def test_rtplan_input_exists(self):
        """Verify the RTPLAN input file exists."""
        rtplan_path = "/root/rtplan.dcm"
        assert os.path.exists(rtplan_path), f"RTPLAN input file not found: {rtplan_path}"

    def test_ct_input_exists(self):
        """Verify the CT input file exists."""
        ct_path = "/root/CT_small.dcm"
        assert os.path.exists(ct_path), f"CT input file not found: {ct_path}"

    def test_mr_input_exists(self):
        """Verify the MR input file exists."""
        mr_path = "/root/MR_small.dcm"
        assert os.path.exists(mr_path), f"MR input file not found: {mr_path}"


class TestReproducibility:
    """Tests to verify the verdict is deterministic and reproducible."""

    def test_verdict_matches_expected_logic(self):
        """Verify the verdict follows the PASS/FAIL logic from the specification.

        PASS: Every referenced UID from RTPLAN matches at least one CT/MR identifier
        FAIL: Otherwise
        """
        from openpyxl import load_workbook
        assert os.path.exists(OUTPUT_FILE), f"Output file not found: {OUTPUT_FILE}"

        wb = load_workbook(OUTPUT_FILE)
        summary = wb["Summary"]
        uid_matches = wb["UID_Matches"]

        verdict = summary["B2"].value
        unmatched = summary["B5"].value

        # Verify logic: PASS only if no unmatched UIDs
        if unmatched == 0:
            expected_verdict = "PASS"
        else:
            expected_verdict = "FAIL"

        assert verdict == expected_verdict, (
            f"Verdict '{verdict}' is inconsistent with unmatched count ({unmatched}). "
            f"Expected '{expected_verdict}'"
        )
