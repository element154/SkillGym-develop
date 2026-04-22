"""Auto-generated expectation tests for FTP command comparison task.

These tests verify that the task execution produces correct outputs
based on the instruction requirements for parsing PCAP files and
creating an Excel workbook with FTP command statistics.
"""

import os
from pathlib import Path
from typing import List, Tuple, Set, Optional

import pytest

# Expected values derived from input file analysis
EXPECTED_OUTPUT_PATH = "/root/ftp_command_comparison.xlsx"
EXPECTED_HTML_TITLE = "SampleCaptures - Wireshark Wiki"

# FTP.pcap expected values (11 commands, each verb appears once)
FTP_PCAP_EXPECTED_VERBS = {"USER", "PASS", "SYST", "FEAT", "CLNT", "CWD", "PWD", "TYPE", "SIZE", "PASV", "RETR"}
FTP_PCAP_TOTAL_COMMANDS = 11
FTP_PCAP_UNIQUE_VERBS = 11
FTP_PCAP_TOP_VERB = "CLNT"  # Alphabetically first when all counts are 1

# Bruteforce.pcap expected values (90 commands: USER=30, PASS=30, QUIT=30)
BRUTEFORCE_PCAP_EXPECTED_VERBS = {"USER", "PASS", "QUIT"}
BRUTEFORCE_PCAP_TOTAL_COMMANDS = 90
BRUTEFORCE_PCAP_UNIQUE_VERBS = 3
BRUTEFORCE_PCAP_TOP_VERB = "PASS"  # Alphabetically first among ties (PASS < QUIT < USER)


def parse_detail_sheet_table(sheet, sheet_name: str) -> Tuple[Optional[Tuple[str, str]], List[Tuple[str, int]], int]:
    """Parse a detail sheet as a contiguous table.

    Returns:
        - header tuple (col_a, col_b) or None if row 1 is empty
        - list of (verb, count) tuples from data rows (row 2 onwards until first blank)
        - first_blank_row: the row number where the table ends (first blank row)
    """
    header_a = sheet.cell(row=1, column=1).value
    header_b = sheet.cell(row=1, column=2).value
    header = (header_a, header_b) if header_a is not None else None

    data = []
    first_blank_row = 2

    for row in range(2, sheet.max_row + 2):  # +2 to check one beyond max_row
        verb = sheet.cell(row=row, column=1).value
        count = sheet.cell(row=row, column=2).value

        if verb is None or (isinstance(verb, str) and verb.strip() == ""):
            first_blank_row = row
            break

        data.append((verb, count))
        first_blank_row = row + 1

    return header, data, first_blank_row


def compute_top_verb(data: List[Tuple[str, int]]) -> Optional[str]:
    """Compute top verb from frequency data (max count, tie-break alphabetically)."""
    if not data:
        return None

    max_count = max(count for _, count in data)
    top_verbs = sorted([verb for verb, count in data if count == max_count])
    return top_verbs[0] if top_verbs else None


class TestOutputFileExists:
    """Tests for output file existence and basic validity."""

    def test_output_file_exists(self):
        """Verify the output Excel file was created at the expected path."""
        assert os.path.exists(EXPECTED_OUTPUT_PATH), (
            f"Output file not found at {EXPECTED_OUTPUT_PATH}"
        )

    def test_output_file_is_not_empty(self):
        """Verify the output file has content."""
        assert os.path.exists(EXPECTED_OUTPUT_PATH), "Output file not found"
        file_size = os.path.getsize(EXPECTED_OUTPUT_PATH)
        assert file_size > 0, "Output file is empty"

    def test_output_file_is_valid_xlsx(self):
        """Verify the output file is a valid Excel workbook."""
        from openpyxl import load_workbook

        assert os.path.exists(EXPECTED_OUTPUT_PATH), "Output file not found"
        try:
            wb = load_workbook(EXPECTED_OUTPUT_PATH)
            assert wb is not None, "Failed to load workbook"
            wb.close()
        except Exception as e:
            pytest.fail(f"Output file is not a valid Excel workbook: {e}")


class TestWorkbookStructure:
    """Tests for the workbook sheet structure."""

    @pytest.fixture
    def workbook(self):
        """Load the output workbook."""
        from openpyxl import load_workbook

        if not os.path.exists(EXPECTED_OUTPUT_PATH):
            pytest.skip("Output file not found")
        wb = load_workbook(EXPECTED_OUTPUT_PATH)
        yield wb
        wb.close()

    def test_has_summary_sheet(self, workbook):
        """Verify the workbook contains a 'Summary' sheet."""
        assert "Summary" in workbook.sheetnames, (
            f"Missing 'Summary' sheet. Found sheets: {workbook.sheetnames}"
        )

    def test_has_ftp_pcap_sheet(self, workbook):
        """Verify the workbook contains a 'ftp_pcap' sheet."""
        assert "ftp_pcap" in workbook.sheetnames, (
            f"Missing 'ftp_pcap' sheet. Found sheets: {workbook.sheetnames}"
        )

    def test_has_bruteforce_pcap_sheet(self, workbook):
        """Verify the workbook contains a 'bruteforce_pcap' sheet."""
        assert "bruteforce_pcap" in workbook.sheetnames, (
            f"Missing 'bruteforce_pcap' sheet. Found sheets: {workbook.sheetnames}"
        )

    def test_has_exactly_three_sheets(self, workbook):
        """Verify the workbook has exactly three required sheets."""
        expected_sheets = {"Summary", "ftp_pcap", "bruteforce_pcap"}
        actual_sheets = set(workbook.sheetnames)
        assert expected_sheets == actual_sheets, (
            f"Expected sheets {expected_sheets}, found {actual_sheets}"
        )


class TestSummarySheet:
    """Tests for the Summary sheet content and structure."""

    @pytest.fixture
    def summary_sheet(self):
        """Load the Summary sheet."""
        from openpyxl import load_workbook

        if not os.path.exists(EXPECTED_OUTPUT_PATH):
            pytest.skip("Output file not found")
        wb = load_workbook(EXPECTED_OUTPUT_PATH)
        if "Summary" not in wb.sheetnames:
            wb.close()
            pytest.skip("Summary sheet not found")
        ws = wb["Summary"]
        yield ws
        wb.close()

    def test_cell_a1_contains_reference_label(self, summary_sheet):
        """Verify cell A1 contains 'Reference Page Title'."""
        value = summary_sheet["A1"].value
        assert value == "Reference Page Title", (
            f"Summary!A1 should be 'Reference Page Title', found '{value}'"
        )

    def test_cell_b1_contains_html_title(self, summary_sheet):
        """Verify cell B1 contains the extracted HTML title."""
        value = summary_sheet["B1"].value
        assert value == EXPECTED_HTML_TITLE, (
            f"Summary!B1 should be '{EXPECTED_HTML_TITLE}', found '{value}'"
        )

    def test_header_row_structure(self, summary_sheet):
        """Verify row 3 contains the expected headers."""
        expected_headers = ["pcap", "total_commands", "unique_verbs", "top_verb"]
        actual_headers = [
            summary_sheet["A3"].value,
            summary_sheet["B3"].value,
            summary_sheet["C3"].value,
            summary_sheet["D3"].value,
        ]
        assert actual_headers == expected_headers, (
            f"Summary row 3 headers should be {expected_headers}, found {actual_headers}"
        )

    def test_ftp_pcap_row_pcap_name(self, summary_sheet):
        """Verify row 4 contains ftp.pcap name."""
        value = summary_sheet["A4"].value
        assert value == "ftp.pcap", f"Summary!A4 should be 'ftp.pcap', found '{value}'"

    def test_ftp_pcap_total_commands(self, summary_sheet):
        """Verify ftp.pcap total commands count."""
        value = summary_sheet["B4"].value
        assert value == FTP_PCAP_TOTAL_COMMANDS, (
            f"Summary!B4 (ftp.pcap total_commands) should be {FTP_PCAP_TOTAL_COMMANDS}, found {value}"
        )

    def test_ftp_pcap_unique_verbs(self, summary_sheet):
        """Verify ftp.pcap unique verbs count."""
        value = summary_sheet["C4"].value
        assert value == FTP_PCAP_UNIQUE_VERBS, (
            f"Summary!C4 (ftp.pcap unique_verbs) should be {FTP_PCAP_UNIQUE_VERBS}, found {value}"
        )

    def test_ftp_pcap_top_verb(self, summary_sheet):
        """Verify ftp.pcap top verb (alphabetically first when tied)."""
        value = summary_sheet["D4"].value
        assert value == FTP_PCAP_TOP_VERB, (
            f"Summary!D4 (ftp.pcap top_verb) should be '{FTP_PCAP_TOP_VERB}', found '{value}'"
        )

    def test_bruteforce_pcap_row_pcap_name(self, summary_sheet):
        """Verify row 5 contains bruteforce.pcap name."""
        value = summary_sheet["A5"].value
        assert value == "bruteforce.pcap", (
            f"Summary!A5 should be 'bruteforce.pcap', found '{value}'"
        )

    def test_bruteforce_pcap_total_commands(self, summary_sheet):
        """Verify bruteforce.pcap total commands count."""
        value = summary_sheet["B5"].value
        assert value == BRUTEFORCE_PCAP_TOTAL_COMMANDS, (
            f"Summary!B5 (bruteforce.pcap total_commands) should be {BRUTEFORCE_PCAP_TOTAL_COMMANDS}, found {value}"
        )

    def test_bruteforce_pcap_unique_verbs(self, summary_sheet):
        """Verify bruteforce.pcap unique verbs count."""
        value = summary_sheet["C5"].value
        assert value == BRUTEFORCE_PCAP_UNIQUE_VERBS, (
            f"Summary!C5 (bruteforce.pcap unique_verbs) should be {BRUTEFORCE_PCAP_UNIQUE_VERBS}, found {value}"
        )

    def test_bruteforce_pcap_top_verb(self, summary_sheet):
        """Verify bruteforce.pcap top verb (alphabetically first among ties)."""
        value = summary_sheet["D5"].value
        assert value == BRUTEFORCE_PCAP_TOP_VERB, (
            f"Summary!D5 (bruteforce.pcap top_verb) should be '{BRUTEFORCE_PCAP_TOP_VERB}', found '{value}'"
        )

    def test_summary_pcap_entries_exactly_two(self, summary_sheet):
        """Verify Summary has exactly two PCAP entries at rows 4 and 5, with no extra entries."""
        pcap_names = [summary_sheet["A4"].value, summary_sheet["A5"].value]
        expected_names = ["ftp.pcap", "bruteforce.pcap"]
        assert pcap_names == expected_names, (
            f"Summary!A4:A5 should be {expected_names}, found {pcap_names}"
        )

        a6_value = summary_sheet["A6"].value
        assert a6_value is None or (isinstance(a6_value, str) and a6_value.strip() == ""), (
            f"Summary!A6 should be empty (no additional PCAP entries), found '{a6_value}'"
        )

        for row in range(7, min(summary_sheet.max_row + 1, 20)):
            cell_value = summary_sheet.cell(row=row, column=1).value
            if cell_value and isinstance(cell_value, str) and ".pcap" in cell_value.lower():
                pytest.fail(
                    f"Summary!A{row} contains unexpected PCAP entry '{cell_value}'. "
                    "Only ftp.pcap and bruteforce.pcap should be present."
                )


class TestFtpPcapSheet:
    """Tests for the ftp_pcap sheet (command frequency table)."""

    @pytest.fixture
    def ftp_sheet(self):
        """Load the ftp_pcap sheet."""
        from openpyxl import load_workbook

        if not os.path.exists(EXPECTED_OUTPUT_PATH):
            pytest.skip("Output file not found")
        wb = load_workbook(EXPECTED_OUTPUT_PATH)
        if "ftp_pcap" not in wb.sheetnames:
            wb.close()
            pytest.skip("ftp_pcap sheet not found")
        ws = wb["ftp_pcap"]
        yield ws
        wb.close()

    @pytest.fixture
    def ftp_parsed_table(self, ftp_sheet):
        """Parse ftp_pcap sheet as a contiguous table."""
        return parse_detail_sheet_table(ftp_sheet, "ftp_pcap")

    def test_ftp_sheet_header_row(self, ftp_sheet):
        """Verify ftp_pcap sheet has correct header row with 'verb' and 'count'."""
        header_a = ftp_sheet.cell(row=1, column=1).value
        header_b = ftp_sheet.cell(row=1, column=2).value

        assert header_a is not None, "ftp_pcap!A1 (header) should not be empty"
        assert header_b is not None, "ftp_pcap!B1 (header) should not be empty"

        assert str(header_a).lower() == "verb", (
            f"ftp_pcap!A1 should be 'verb', found '{header_a}'"
        )
        assert str(header_b).lower() == "count", (
            f"ftp_pcap!B1 should be 'count', found '{header_b}'"
        )

    def test_ftp_sheet_has_correct_row_count(self, ftp_parsed_table):
        """Verify ftp_pcap sheet has exactly the expected number of data rows."""
        header, data, first_blank = ftp_parsed_table

        assert len(data) == FTP_PCAP_UNIQUE_VERBS, (
            f"ftp_pcap should have exactly {FTP_PCAP_UNIQUE_VERBS} data rows, "
            f"found {len(data)} rows before first blank at row {first_blank}"
        )

    def test_ftp_sheet_contains_exact_verb_set(self, ftp_parsed_table):
        """Verify ftp_pcap sheet contains exactly the expected FTP verbs (no more, no less)."""
        header, data, first_blank = ftp_parsed_table

        verbs_found = {verb.upper() if isinstance(verb, str) else verb for verb, _ in data}

        missing_verbs = FTP_PCAP_EXPECTED_VERBS - verbs_found
        extra_verbs = verbs_found - FTP_PCAP_EXPECTED_VERBS

        assert not missing_verbs, (
            f"ftp_pcap missing expected verbs: {missing_verbs}. Found: {verbs_found}"
        )
        assert not extra_verbs, (
            f"ftp_pcap contains unexpected verbs: {extra_verbs}. Expected: {FTP_PCAP_EXPECTED_VERBS}"
        )

    def test_ftp_sheet_no_duplicate_verbs(self, ftp_parsed_table):
        """Verify ftp_pcap sheet has no duplicate verb entries."""
        header, data, first_blank = ftp_parsed_table

        verbs = [verb.upper() if isinstance(verb, str) else verb for verb, _ in data]
        seen = set()
        duplicates = []

        for i, verb in enumerate(verbs, start=2):
            if verb in seen:
                duplicates.append(f"'{verb}' at row {i}")
            seen.add(verb)

        assert not duplicates, (
            f"ftp_pcap contains duplicate verbs: {', '.join(duplicates)}"
        )

    def test_ftp_sheet_all_counts_are_positive_integers(self, ftp_parsed_table):
        """Verify all verb counts in ftp_pcap are positive integers."""
        header, data, first_blank = ftp_parsed_table

        for i, (verb, count) in enumerate(data, start=2):
            assert isinstance(count, int), (
                f"ftp_pcap!B{i} (count for '{verb}') should be an integer, "
                f"found {type(count).__name__}: {count}"
            )
            assert count > 0, (
                f"ftp_pcap!B{i} (count for '{verb}') should be positive, found {count}"
            )

    def test_ftp_sheet_all_counts_are_one(self, ftp_parsed_table):
        """Verify all verb counts in ftp_pcap are 1 (each appears once)."""
        header, data, first_blank = ftp_parsed_table

        for i, (verb, count) in enumerate(data, start=2):
            assert count == 1, (
                f"ftp_pcap!B{i} (count for '{verb}') should be 1, found {count}"
            )

    def test_ftp_sheet_sorted_correctly(self, ftp_parsed_table):
        """Verify ftp_pcap sheet is sorted by count desc, then verb asc."""
        header, data, first_blank = ftp_parsed_table

        if len(data) <= 1:
            return

        normalized_data = [
            (verb.upper() if isinstance(verb, str) else verb, count)
            for verb, count in data
        ]

        expected_sorted = sorted(normalized_data, key=lambda x: (-x[1], x[0]))

        assert normalized_data == expected_sorted, (
            f"ftp_pcap data rows not sorted correctly by (count DESC, verb ASC).\n"
            f"Expected order: {[v for v, c in expected_sorted]}\n"
            f"Actual order:   {[v for v, c in normalized_data]}"
        )

    def test_ftp_sheet_table_is_contiguous(self, ftp_sheet, ftp_parsed_table):
        """Verify ftp_pcap table is contiguous with no gaps or extra rows after."""
        header, data, first_blank = ftp_parsed_table

        expected_first_blank = 2 + FTP_PCAP_UNIQUE_VERBS
        assert first_blank == expected_first_blank, (
            f"ftp_pcap table should end at row {expected_first_blank}, "
            f"but first blank/end found at row {first_blank}"
        )

        for row in range(first_blank, min(ftp_sheet.max_row + 1, first_blank + 10)):
            cell_a = ftp_sheet.cell(row=row, column=1).value
            cell_b = ftp_sheet.cell(row=row, column=2).value
            if cell_a is not None and str(cell_a).strip() != "":
                pytest.fail(
                    f"ftp_pcap!A{row} contains unexpected data '{cell_a}' after table end. "
                    "Table should be contiguous."
                )


class TestBruteforcePcapSheet:
    """Tests for the bruteforce_pcap sheet (command frequency table)."""

    @pytest.fixture
    def bruteforce_sheet(self):
        """Load the bruteforce_pcap sheet."""
        from openpyxl import load_workbook

        if not os.path.exists(EXPECTED_OUTPUT_PATH):
            pytest.skip("Output file not found")
        wb = load_workbook(EXPECTED_OUTPUT_PATH)
        if "bruteforce_pcap" not in wb.sheetnames:
            wb.close()
            pytest.skip("bruteforce_pcap sheet not found")
        ws = wb["bruteforce_pcap"]
        yield ws
        wb.close()

    @pytest.fixture
    def bruteforce_parsed_table(self, bruteforce_sheet):
        """Parse bruteforce_pcap sheet as a contiguous table."""
        return parse_detail_sheet_table(bruteforce_sheet, "bruteforce_pcap")

    def test_bruteforce_sheet_header_row(self, bruteforce_sheet):
        """Verify bruteforce_pcap sheet has correct header row with 'verb' and 'count'."""
        header_a = bruteforce_sheet.cell(row=1, column=1).value
        header_b = bruteforce_sheet.cell(row=1, column=2).value

        assert header_a is not None, "bruteforce_pcap!A1 (header) should not be empty"
        assert header_b is not None, "bruteforce_pcap!B1 (header) should not be empty"

        assert str(header_a).lower() == "verb", (
            f"bruteforce_pcap!A1 should be 'verb', found '{header_a}'"
        )
        assert str(header_b).lower() == "count", (
            f"bruteforce_pcap!B1 should be 'count', found '{header_b}'"
        )

    def test_bruteforce_sheet_has_correct_row_count(self, bruteforce_parsed_table):
        """Verify bruteforce_pcap sheet has exactly the expected number of data rows."""
        header, data, first_blank = bruteforce_parsed_table

        assert len(data) == BRUTEFORCE_PCAP_UNIQUE_VERBS, (
            f"bruteforce_pcap should have exactly {BRUTEFORCE_PCAP_UNIQUE_VERBS} data rows, "
            f"found {len(data)} rows before first blank at row {first_blank}"
        )

    def test_bruteforce_sheet_contains_exact_verb_set(self, bruteforce_parsed_table):
        """Verify bruteforce_pcap sheet contains exactly the expected FTP verbs."""
        header, data, first_blank = bruteforce_parsed_table

        verbs_found = {verb.upper() if isinstance(verb, str) else verb for verb, _ in data}

        missing_verbs = BRUTEFORCE_PCAP_EXPECTED_VERBS - verbs_found
        extra_verbs = verbs_found - BRUTEFORCE_PCAP_EXPECTED_VERBS

        assert not missing_verbs, (
            f"bruteforce_pcap missing expected verbs: {missing_verbs}. Found: {verbs_found}"
        )
        assert not extra_verbs, (
            f"bruteforce_pcap contains unexpected verbs: {extra_verbs}. "
            f"Expected: {BRUTEFORCE_PCAP_EXPECTED_VERBS}"
        )

    def test_bruteforce_sheet_no_duplicate_verbs(self, bruteforce_parsed_table):
        """Verify bruteforce_pcap sheet has no duplicate verb entries."""
        header, data, first_blank = bruteforce_parsed_table

        verbs = [verb.upper() if isinstance(verb, str) else verb for verb, _ in data]
        seen = set()
        duplicates = []

        for i, verb in enumerate(verbs, start=2):
            if verb in seen:
                duplicates.append(f"'{verb}' at row {i}")
            seen.add(verb)

        assert not duplicates, (
            f"bruteforce_pcap contains duplicate verbs: {', '.join(duplicates)}"
        )

    def test_bruteforce_sheet_all_counts_are_positive_integers(self, bruteforce_parsed_table):
        """Verify all verb counts in bruteforce_pcap are positive integers."""
        header, data, first_blank = bruteforce_parsed_table

        for i, (verb, count) in enumerate(data, start=2):
            assert isinstance(count, int), (
                f"bruteforce_pcap!B{i} (count for '{verb}') should be an integer, "
                f"found {type(count).__name__}: {count}"
            )
            assert count > 0, (
                f"bruteforce_pcap!B{i} (count for '{verb}') should be positive, found {count}"
            )

    def test_bruteforce_sheet_correct_counts(self, bruteforce_parsed_table):
        """Verify verb counts in bruteforce_pcap are correct (30 each)."""
        header, data, first_blank = bruteforce_parsed_table
        expected_counts = {"USER": 30, "PASS": 30, "QUIT": 30}

        for i, (verb, count) in enumerate(data, start=2):
            verb_upper = verb.upper() if isinstance(verb, str) else verb
            if verb_upper in expected_counts:
                expected = expected_counts[verb_upper]
                assert count == expected, (
                    f"bruteforce_pcap!B{i} (count for '{verb}') should be {expected}, found {count}"
                )

    def test_bruteforce_sheet_sorted_correctly(self, bruteforce_parsed_table):
        """Verify bruteforce_pcap sheet is sorted by count desc, then verb asc."""
        header, data, first_blank = bruteforce_parsed_table

        if len(data) <= 1:
            return

        normalized_data = [
            (verb.upper() if isinstance(verb, str) else verb, count)
            for verb, count in data
        ]

        expected_sorted = sorted(normalized_data, key=lambda x: (-x[1], x[0]))

        assert normalized_data == expected_sorted, (
            f"bruteforce_pcap data rows not sorted correctly by (count DESC, verb ASC).\n"
            f"Expected order: {[v for v, c in expected_sorted]}\n"
            f"Actual order:   {[v for v, c in normalized_data]}"
        )

    def test_bruteforce_sheet_table_is_contiguous(self, bruteforce_sheet, bruteforce_parsed_table):
        """Verify bruteforce_pcap table is contiguous with no gaps or extra rows after."""
        header, data, first_blank = bruteforce_parsed_table

        expected_first_blank = 2 + BRUTEFORCE_PCAP_UNIQUE_VERBS
        assert first_blank == expected_first_blank, (
            f"bruteforce_pcap table should end at row {expected_first_blank}, "
            f"but first blank/end found at row {first_blank}"
        )

        for row in range(first_blank, min(bruteforce_sheet.max_row + 1, first_blank + 10)):
            cell_a = bruteforce_sheet.cell(row=row, column=1).value
            if cell_a is not None and str(cell_a).strip() != "":
                pytest.fail(
                    f"bruteforce_pcap!A{row} contains unexpected data '{cell_a}' after table end. "
                    "Table should be contiguous."
                )


class TestDataIntegrity:
    """Tests for overall data integrity and consistency."""

    @pytest.fixture
    def workbook(self):
        """Load the output workbook."""
        from openpyxl import load_workbook

        if not os.path.exists(EXPECTED_OUTPUT_PATH):
            pytest.skip("Output file not found")
        wb = load_workbook(EXPECTED_OUTPUT_PATH)
        yield wb
        wb.close()

    def test_ftp_unique_verbs_matches_detail_sheet(self, workbook):
        """Verify Summary ftp.pcap unique_verbs equals distinct verbs in ftp_pcap sheet."""
        if "Summary" not in workbook.sheetnames:
            pytest.skip("Summary sheet not found")
        if "ftp_pcap" not in workbook.sheetnames:
            pytest.skip("ftp_pcap sheet not found")

        summary = workbook["Summary"]
        ftp_sheet = workbook["ftp_pcap"]

        header, data, first_blank = parse_detail_sheet_table(ftp_sheet, "ftp_pcap")

        verbs_in_sheet = {
            verb.upper() if isinstance(verb, str) else verb
            for verb, _ in data
        }

        summary_unique = summary["C4"].value

        assert len(verbs_in_sheet) == summary_unique, (
            f"ftp_pcap distinct verb count ({len(verbs_in_sheet)}) doesn't match "
            f"Summary!C4 unique_verbs ({summary_unique}). Verbs found: {verbs_in_sheet}"
        )

        assert verbs_in_sheet == FTP_PCAP_EXPECTED_VERBS, (
            f"ftp_pcap verb set {verbs_in_sheet} doesn't match expected {FTP_PCAP_EXPECTED_VERBS}"
        )

    def test_bruteforce_unique_verbs_matches_detail_sheet(self, workbook):
        """Verify Summary bruteforce.pcap unique_verbs equals distinct verbs in bruteforce_pcap sheet."""
        if "Summary" not in workbook.sheetnames:
            pytest.skip("Summary sheet not found")
        if "bruteforce_pcap" not in workbook.sheetnames:
            pytest.skip("bruteforce_pcap sheet not found")

        summary = workbook["Summary"]
        bf_sheet = workbook["bruteforce_pcap"]

        header, data, first_blank = parse_detail_sheet_table(bf_sheet, "bruteforce_pcap")

        verbs_in_sheet = {
            verb.upper() if isinstance(verb, str) else verb
            for verb, _ in data
        }

        summary_unique = summary["C5"].value

        assert len(verbs_in_sheet) == summary_unique, (
            f"bruteforce_pcap distinct verb count ({len(verbs_in_sheet)}) doesn't match "
            f"Summary!C5 unique_verbs ({summary_unique}). Verbs found: {verbs_in_sheet}"
        )

        assert verbs_in_sheet == BRUTEFORCE_PCAP_EXPECTED_VERBS, (
            f"bruteforce_pcap verb set {verbs_in_sheet} doesn't match expected {BRUTEFORCE_PCAP_EXPECTED_VERBS}"
        )

    def test_ftp_pcap_total_matches_sum_of_counts(self, workbook):
        """Verify ftp.pcap total_commands equals sum of all verb counts from detail sheet."""
        if "Summary" not in workbook.sheetnames or "ftp_pcap" not in workbook.sheetnames:
            pytest.skip("Required sheets not found")

        summary = workbook["Summary"]
        ftp_sheet = workbook["ftp_pcap"]

        header, data, first_blank = parse_detail_sheet_table(ftp_sheet, "ftp_pcap")

        total_from_summary = summary["B4"].value
        total_from_counts = sum(count for _, count in data if isinstance(count, int))

        assert total_from_summary == total_from_counts, (
            f"Summary!B4 ftp.pcap total_commands ({total_from_summary}) doesn't match "
            f"sum of ftp_pcap counts ({total_from_counts})"
        )

    def test_bruteforce_pcap_total_matches_sum_of_counts(self, workbook):
        """Verify bruteforce.pcap total_commands equals sum of all verb counts from detail sheet."""
        if "Summary" not in workbook.sheetnames or "bruteforce_pcap" not in workbook.sheetnames:
            pytest.skip("Required sheets not found")

        summary = workbook["Summary"]
        bf_sheet = workbook["bruteforce_pcap"]

        header, data, first_blank = parse_detail_sheet_table(bf_sheet, "bruteforce_pcap")

        total_from_summary = summary["B5"].value
        total_from_counts = sum(count for _, count in data if isinstance(count, int))

        assert total_from_summary == total_from_counts, (
            f"Summary!B5 bruteforce.pcap total_commands ({total_from_summary}) doesn't match "
            f"sum of bruteforce_pcap counts ({total_from_counts})"
        )

    def test_ftp_top_verb_derived_from_detail_sheet(self, workbook):
        """Verify Summary ftp.pcap top_verb matches computation from ftp_pcap sheet."""
        if "Summary" not in workbook.sheetnames or "ftp_pcap" not in workbook.sheetnames:
            pytest.skip("Required sheets not found")

        summary = workbook["Summary"]
        ftp_sheet = workbook["ftp_pcap"]

        header, data, first_blank = parse_detail_sheet_table(ftp_sheet, "ftp_pcap")

        normalized_data = [
            (verb.upper() if isinstance(verb, str) else verb, count)
            for verb, count in data
            if isinstance(count, int)
        ]

        computed_top_verb = compute_top_verb(normalized_data)
        summary_top_verb = summary["D4"].value

        assert summary_top_verb == computed_top_verb, (
            f"Summary!D4 ftp.pcap top_verb ('{summary_top_verb}') doesn't match "
            f"computed top verb from ftp_pcap sheet ('{computed_top_verb}'). "
            f"Top verb should be max count, tie-break alphabetically."
        )

    def test_bruteforce_top_verb_derived_from_detail_sheet(self, workbook):
        """Verify Summary bruteforce.pcap top_verb matches computation from bruteforce_pcap sheet."""
        if "Summary" not in workbook.sheetnames or "bruteforce_pcap" not in workbook.sheetnames:
            pytest.skip("Required sheets not found")

        summary = workbook["Summary"]
        bf_sheet = workbook["bruteforce_pcap"]

        header, data, first_blank = parse_detail_sheet_table(bf_sheet, "bruteforce_pcap")

        normalized_data = [
            (verb.upper() if isinstance(verb, str) else verb, count)
            for verb, count in data
            if isinstance(count, int)
        ]

        computed_top_verb = compute_top_verb(normalized_data)
        summary_top_verb = summary["D5"].value

        assert summary_top_verb == computed_top_verb, (
            f"Summary!D5 bruteforce.pcap top_verb ('{summary_top_verb}') doesn't match "
            f"computed top verb from bruteforce_pcap sheet ('{computed_top_verb}'). "
            f"Top verb should be max count, tie-break alphabetically."
        )
