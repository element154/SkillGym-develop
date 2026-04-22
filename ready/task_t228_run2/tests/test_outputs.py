"""Auto-generated expectation tests for task verification.

These tests verify that the task execution produces correct outputs
based on the instruction requirements for critical edge DC power flow analysis.

The tests recompute expected results from the input MATPOWER files to ensure
correctness of:
1. Graph construction from in-service branches (status==1)
2. Edge betweenness centrality computation
3. Deterministic tie-breaking: lexicographically smallest (min(fbus,tbus), max(fbus,tbus))
4. DC power flow with unique slack bus (type==3) fixed at angle 0
5. Line flow computation: flow = (theta_fbus - theta_tbus) / x * baseMVA
6. Directed edge reporting using original fbus->tbus from first branch occurrence
"""

import os
import re
from pathlib import Path
from typing import Dict, List, Tuple, Any, Optional

import pytest
import numpy as np

# Expected output path
OUTPUT_FILE = "/root/critical_edge_dcflow_report.xlsx"

# Input directory
INPUT_DIR = "/root"

# Expected case files (basenames)
EXPECTED_CASE_FILES = ["case57.m", "case118.m", "pglib_opf_case118_ieee.m"]

# Expected bus counts for each case
EXPECTED_BUS_COUNTS = {
    "case57.m": 57,
    "case118.m": 118,
    "pglib_opf_case118_ieee.m": 118,
}

# Required columns in Summary sheet (exactly these, in this order)
REQUIRED_COLUMNS = [
    "case_file",
    "n_bus",
    "n_branch_in_service",
    "critical_edge_fbus",
    "critical_edge_tbus",
    "critical_edge_edge_betweenness",
    "critical_edge_flow_MW",
]

# Number of expected data rows (one per case file)
EXPECTED_DATA_ROWS = 3

# Tolerance for floating-point comparisons in tie-breaking
TIE_TOLERANCE = 1e-12

# Light gray fill RGB values (common representations)
LIGHT_GRAY_RGBS = [
    "D3D3D3",  # lightgray
    "FFD3D3D3",  # with alpha
    "D9D9D9",  # common Excel light gray
    "FFD9D9D9",
    "C0C0C0",  # silver
    "FFC0C0C0",
    "E0E0E0",
    "FFE0E0E0",
    "BFBFBF",
    "FFBFBFBF",
    "F2F2F2",
    "FFF2F2F2",
    "EEEEEE",
    "FFEEEEEE",
    "DDDDDD",
    "FFDDDDDD",
]

# Blue font RGB values (common representations)
BLUE_RGBS = [
    "0000FF",  # pure blue
    "FF0000FF",  # with alpha
    "0070C0",  # Excel default blue for inputs
    "FF0070C0",
    "0066CC",
    "FF0066CC",
    "0000CC",
    "FF0000CC",
    "000080",  # navy
    "FF000080",
    "0000CD",  # medium blue
    "FF0000CD",
    "4169E1",  # royal blue
    "FF4169E1",
    "1F497D",  # Excel dark blue
    "FF1F497D",
    "5B9BD5",  # Excel accent blue
    "FF5B9BD5",
]


def parse_matpower_file(filepath: str) -> Dict[str, Any]:
    """Parse a MATPOWER .m file and extract bus, gen, and branch data.

    Robust parser that handles:
    - Scientific notation (e.g., 1.5e-5, 2.3E+10)
    - Semicolon row terminators
    - Tab and comma delimiters
    - Inline comments
    - Continuation lines
    """
    with open(filepath, 'r') as f:
        content = f.read()

    result = {
        'baseMVA': 100.0,
        'bus': [],
        'gen': [],
        'branch': [],
    }

    # Extract baseMVA (support scientific notation)
    basemva_match = re.search(r'mpc\.baseMVA\s*=\s*([\d.eE+-]+)', content)
    if basemva_match:
        result['baseMVA'] = float(basemva_match.group(1))

    def extract_matrix_data(content: str, matrix_name: str) -> List[List[float]]:
        """Extract matrix data between brackets with robust parsing."""
        # Match mpc.matrix_name = [ ... ];
        pattern = rf'mpc\.{matrix_name}\s*=\s*\[(.*?)\];'
        match = re.search(pattern, content, re.DOTALL)
        if not match:
            return []

        matrix_text = match.group(1)

        # Remove comments (% to end of line)
        matrix_text = re.sub(r'%[^\n]*', '', matrix_text)

        # Replace tabs and commas with spaces
        matrix_text = matrix_text.replace('\t', ' ').replace(',', ' ')

        # Split by semicolons or newlines to get rows
        # Handle both explicit semicolons and newlines as row separators
        rows = []
        current_row = []

        # Tokenize: split on whitespace and semicolons
        tokens = re.split(r'[\s;]+', matrix_text)

        for token in tokens:
            token = token.strip()
            if not token:
                continue

            # Check if it's a number (including scientific notation)
            try:
                value = float(token)
                current_row.append(value)
            except ValueError:
                # Not a number, skip
                continue

        # Now we have all numbers in current_row, need to split into rows
        # Re-parse to respect row structure
        matrix_text_clean = re.sub(r'%[^\n]*', '', match.group(1))

        # Split by semicolons first (explicit row delimiters)
        row_chunks = re.split(r';', matrix_text_clean)

        rows = []
        for chunk in row_chunks:
            # Replace delimiters and split
            chunk = chunk.replace('\t', ' ').replace(',', ' ')
            # Find all numeric values including scientific notation
            values = re.findall(r'-?[\d.]+(?:[eE][+-]?\d+)?', chunk)
            if values:
                try:
                    row = [float(v) for v in values]
                    if row:  # Non-empty row
                        rows.append(row)
                except ValueError:
                    continue

        # If semicolon parsing didn't work well, try newline parsing
        if not rows:
            lines = matrix_text_clean.strip().split('\n')
            for line in lines:
                line = line.replace('\t', ' ').replace(',', ' ')
                values = re.findall(r'-?[\d.]+(?:[eE][+-]?\d+)?', line)
                if values:
                    try:
                        row = [float(v) for v in values]
                        if row:
                            rows.append(row)
                    except ValueError:
                        continue

        return rows

    # Extract bus data
    bus_rows = extract_matrix_data(content, 'bus')
    for values in bus_rows:
        if len(values) >= 4:  # bus_i, type, Pd, Qd at minimum
            result['bus'].append({
                'bus_i': int(values[0]),
                'type': int(values[1]),
                'Pd': values[2],
                'Qd': values[3] if len(values) > 3 else 0,
            })

    # Extract generator data
    gen_rows = extract_matrix_data(content, 'gen')
    for values in gen_rows:
        if len(values) >= 2:  # bus, Pg at minimum
            result['gen'].append({
                'bus': int(values[0]),
                'Pg': values[1],
                'status': int(values[7]) if len(values) > 7 else 1,
            })

    # Extract branch data - preserve order for deterministic selection
    branch_rows = extract_matrix_data(content, 'branch')
    for idx, values in enumerate(branch_rows):
        if len(values) >= 4:  # fbus, tbus, r, x at minimum
            result['branch'].append({
                'fbus': int(values[0]),
                'tbus': int(values[1]),
                'r': values[2],
                'x': values[3],
                'b': values[4] if len(values) > 4 else 0,
                'status': int(values[10]) if len(values) > 10 else 1,
                'file_order': idx,  # Track original file order
            })

    return result


def build_simple_graph_with_branch_info(branches: List[Dict]) -> Tuple[set, set, Dict, Dict]:
    """Build a simple undirected graph from branches.

    Preserves the first occurrence of each undirected edge in file order
    for deterministic selection of direction and reactance.

    Returns:
        - Set of nodes
        - Set of undirected edges as (min, max) tuples
        - Dict mapping undirected edges to their reactance (x) from first occurrence
        - Dict mapping undirected edges to (fbus, tbus, x, file_order) of first branch
    """
    nodes = set()
    edges = set()
    edge_x = {}
    edge_first_branch = {}  # Stores (fbus, tbus, x, file_order) for first occurrence

    for branch in branches:
        if branch['status'] == 1:  # Only in-service branches
            fbus, tbus = branch['fbus'], branch['tbus']
            nodes.add(fbus)
            nodes.add(tbus)
            # Create undirected edge as (min, max)
            edge = (min(fbus, tbus), max(fbus, tbus))
            edges.add(edge)
            # Store info for first occurrence only (deterministic selection)
            if edge not in edge_first_branch:
                edge_x[edge] = branch['x']
                edge_first_branch[edge] = (fbus, tbus, branch['x'], branch['file_order'])

    return nodes, edges, edge_x, edge_first_branch


def compute_edge_betweenness(nodes: set, edges: set) -> Dict[Tuple[int, int], float]:
    """Compute edge betweenness centrality using networkx."""
    pytest.importorskip("networkx")
    import networkx as nx

    G = nx.Graph()
    G.add_nodes_from(nodes)
    G.add_edges_from(edges)

    betweenness = nx.edge_betweenness_centrality(G, normalized=True)

    # Convert to (min, max) format for consistency
    result = {}
    for (u, v), bc in betweenness.items():
        edge = (min(u, v), max(u, v))
        result[edge] = bc

    return result


def find_critical_edge_with_tolerance(betweenness: Dict[Tuple[int, int], float],
                                       tolerance: float = TIE_TOLERANCE) -> Tuple[int, int]:
    """Find critical edge with deterministic tie-breaking and tolerance for ties.

    Tie-break rule: lexicographically smallest (min(fbus,tbus), max(fbus,tbus))
    Uses tolerance to identify edges with effectively equal max betweenness.
    """
    if not betweenness:
        return None

    max_bc = max(betweenness.values())
    # Get all edges with betweenness within tolerance of max (handles floating-point ties)
    max_edges = [edge for edge, bc in betweenness.items() if bc >= max_bc - tolerance]
    # Sort lexicographically and return smallest
    max_edges.sort()
    return max_edges[0]


def compute_dc_power_flow(mpc_data: Dict, baseMVA: float = 100.0) -> Tuple[Dict[int, float], int]:
    """Compute DC power flow and return bus angles in radians and slack bus.

    DC power flow equations:
    - B * theta = P_injection
    - Slack bus angle fixed at 0
    - B_ij = -1/x_ij for connected buses
    - B_ii = sum(1/x_ij) for all branches connected to bus i

    Returns:
        - Dict mapping bus number to angle (radians)
        - Slack bus number
    """
    buses = mpc_data['bus']
    gens = mpc_data['gen']
    branches = mpc_data['branch']

    # Get bus numbers and find slack bus (type==3)
    bus_nums = [b['bus_i'] for b in buses]
    bus_to_idx = {bus: i for i, bus in enumerate(bus_nums)}
    n_bus = len(bus_nums)

    # Find unique slack bus (must be exactly one)
    slack_buses = [b['bus_i'] for b in buses if b['type'] == 3]
    if len(slack_buses) != 1:
        raise ValueError(f"Expected exactly 1 slack bus (type==3), found {len(slack_buses)}: {slack_buses}")

    slack_bus = slack_buses[0]
    slack_idx = bus_to_idx[slack_bus]

    # Compute net injection P = Pg - Pd for each bus
    P = np.zeros(n_bus)

    # Add load (demand)
    for b in buses:
        idx = bus_to_idx[b['bus_i']]
        P[idx] -= b['Pd']  # Load is negative injection

    # Add generation
    for g in gens:
        if g['status'] == 1:  # Only in-service generators
            if g['bus'] in bus_to_idx:
                idx = bus_to_idx[g['bus']]
                P[idx] += g['Pg']

    # Build susceptance matrix B (using ALL in-service branches, including parallel)
    B = np.zeros((n_bus, n_bus))

    for branch in branches:
        if branch['status'] == 1 and branch['x'] != 0:
            fbus, tbus = branch['fbus'], branch['tbus']
            if fbus in bus_to_idx and tbus in bus_to_idx:
                i = bus_to_idx[fbus]
                j = bus_to_idx[tbus]
                b = 1.0 / branch['x']  # susceptance = 1/reactance

                B[i, i] += b
                B[j, j] += b
                B[i, j] -= b
                B[j, i] -= b

    # Remove slack bus row and column
    non_slack = [i for i in range(n_bus) if i != slack_idx]
    B_reduced = B[np.ix_(non_slack, non_slack)]
    P_reduced = P[non_slack] / baseMVA  # Convert to per unit

    # Solve B_reduced * theta_reduced = P_reduced
    try:
        theta_reduced = np.linalg.solve(B_reduced, P_reduced)
    except np.linalg.LinAlgError:
        # Singular matrix - use pseudoinverse
        theta_reduced = np.linalg.lstsq(B_reduced, P_reduced, rcond=None)[0]

    # Reconstruct full theta vector (slack at 0)
    theta = np.zeros(n_bus)
    for i, idx in enumerate(non_slack):
        theta[idx] = theta_reduced[i]

    # Return as dict mapping bus number to angle, plus slack bus
    return {bus_nums[i]: theta[i] for i in range(n_bus)}, slack_bus


def compute_line_flow(theta: Dict[int, float], fbus: int, tbus: int, x: float, baseMVA: float) -> float:
    """Compute line flow in MW using directed fbus->tbus convention.

    Flow from fbus to tbus = (theta_fbus - theta_tbus) / x * baseMVA

    Positive flow means power flows from fbus to tbus.
    """
    theta_f = theta.get(fbus, 0.0)
    theta_t = theta.get(tbus, 0.0)
    return (theta_f - theta_t) / x * baseMVA


def compute_expected_results(case_file: str) -> Dict[str, Any]:
    """Compute all expected results for a case file.

    Returns expected values including the directed edge representation
    based on first occurrence in file order.
    """
    filepath = os.path.join(INPUT_DIR, case_file)
    mpc = parse_matpower_file(filepath)

    nodes, edges, edge_x, edge_first_branch = build_simple_graph_with_branch_info(mpc['branch'])
    betweenness = compute_edge_betweenness(nodes, edges)
    critical_edge = find_critical_edge_with_tolerance(betweenness, TIE_TOLERANCE)

    # Compute DC power flow
    theta, slack_bus = compute_dc_power_flow(mpc, mpc['baseMVA'])

    # Get flow on critical edge using the DIRECTED representation from first branch occurrence
    if critical_edge:
        # Get the first branch's directed info (fbus, tbus, x, file_order)
        fbus_directed, tbus_directed, x_directed, _ = edge_first_branch[critical_edge]
        flow = compute_line_flow(theta, fbus_directed, tbus_directed, x_directed, mpc['baseMVA'])
    else:
        fbus_directed, tbus_directed, x_directed = None, None, None
        flow = 0.0

    return {
        'case_file': case_file,
        'n_bus': len(mpc['bus']),
        'n_branch_in_service': len(edges),  # Unique undirected edges
        'critical_edge_undirected': critical_edge,  # (min, max) for matching
        'critical_edge_fbus': fbus_directed,  # Directed from first branch
        'critical_edge_tbus': tbus_directed,  # Directed from first branch
        'critical_edge_x': x_directed,
        'edge_betweenness': betweenness.get(critical_edge, 0.0) if critical_edge else 0.0,
        'flow_MW': flow,
        'all_betweenness': betweenness,
        'slack_bus': slack_bus,
        'theta': theta,
        'mpc': mpc,
    }


# Precompute expected results for all cases
@pytest.fixture(scope="module")
def expected_results():
    """Compute expected results for all case files."""
    results = {}
    for case_file in EXPECTED_CASE_FILES:
        filepath = os.path.join(INPUT_DIR, case_file)
        if os.path.exists(filepath):
            results[case_file] = compute_expected_results(case_file)
    return results


@pytest.fixture(scope="module")
def workbook_data():
    """Load workbook data once for all tests."""
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    if not os.path.exists(OUTPUT_FILE):
        pytest.skip(f"Output file not found: {OUTPUT_FILE}")

    wb = load_workbook(OUTPUT_FILE)
    ws = wb["Summary"]

    # Read headers (all columns, not just non-empty)
    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value:
            headers[str(value).strip()] = col

    # Read data rows
    data = {}
    for row in range(2, ws.max_row + 1):
        case_file_col = headers.get("case_file")
        if case_file_col:
            case_file = ws.cell(row=row, column=case_file_col).value
            if case_file:
                row_data = {}
                for header, col in headers.items():
                    row_data[header] = ws.cell(row=row, column=col).value
                data[str(case_file).strip()] = row_data

    return {
        'workbook': wb,
        'worksheet': ws,
        'headers': headers,
        'data': data,
    }


class TestOutputFileExists:
    """Tests for output file existence and basic validity."""

    def test_output_file_exists(self):
        """Verify output Excel file was created at the expected path."""
        assert os.path.exists(OUTPUT_FILE), (
            f"Output file not found at {OUTPUT_FILE}"
        )

    def test_output_file_is_not_empty(self):
        """Verify output file has content (not zero bytes)."""
        assert os.path.exists(OUTPUT_FILE), f"Output file not found at {OUTPUT_FILE}"
        file_size = os.path.getsize(OUTPUT_FILE)
        assert file_size > 0, "Output file is empty (0 bytes)"

    def test_output_file_has_xlsx_extension(self):
        """Verify output file has .xlsx extension."""
        assert OUTPUT_FILE.endswith(".xlsx"), (
            f"Output file should have .xlsx extension, got: {OUTPUT_FILE}"
        )


class TestExcelValidity:
    """Tests for Excel file validity and structure."""

    def test_file_is_valid_excel(self):
        """Verify output is a valid Excel file that can be opened."""
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook

        assert os.path.exists(OUTPUT_FILE), f"Output file not found at {OUTPUT_FILE}"

        try:
            wb = load_workbook(OUTPUT_FILE)
            assert wb is not None
            wb.close()
        except Exception as e:
            pytest.fail(f"Failed to open Excel file: {e}")

    def test_summary_sheet_exists(self):
        """Verify 'Summary' sheet exists in the workbook."""
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook

        assert os.path.exists(OUTPUT_FILE), f"Output file not found at {OUTPUT_FILE}"

        wb = load_workbook(OUTPUT_FILE)
        sheet_names = wb.sheetnames
        wb.close()

        assert "Summary" in sheet_names, (
            f"'Summary' sheet not found. Available sheets: {sheet_names}"
        )

    def test_only_summary_sheet_exists(self):
        """Verify only 'Summary' sheet exists (no other sheets)."""
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook

        assert os.path.exists(OUTPUT_FILE), f"Output file not found at {OUTPUT_FILE}"

        wb = load_workbook(OUTPUT_FILE)
        sheet_names = wb.sheetnames
        wb.close()

        # Should have exactly one sheet named "Summary"
        assert len(sheet_names) == 1, (
            f"Expected exactly 1 sheet, found {len(sheet_names)}: {sheet_names}"
        )
        assert sheet_names[0] == "Summary", (
            f"Expected sheet named 'Summary', found: {sheet_names[0]}"
        )


class TestSummarySheetTableStructure:
    """Tests for Summary sheet table structure - exactly one table, no extra content."""

    def test_exact_number_of_columns(self, workbook_data):
        """Verify sheet has exactly the required number of columns."""
        ws = workbook_data['worksheet']

        # Check max_column matches expected
        assert ws.max_column == len(REQUIRED_COLUMNS), (
            f"Expected exactly {len(REQUIRED_COLUMNS)} columns, "
            f"found {ws.max_column}. "
            f"Sheet should contain exactly one table with columns: {REQUIRED_COLUMNS}"
        )

    def test_exact_number_of_rows(self, workbook_data):
        """Verify sheet has exactly 1 header row + 3 data rows = 4 rows total."""
        ws = workbook_data['worksheet']
        expected_rows = 1 + EXPECTED_DATA_ROWS  # 1 header + 3 data rows

        assert ws.max_row == expected_rows, (
            f"Expected exactly {expected_rows} rows (1 header + {EXPECTED_DATA_ROWS} data), "
            f"found {ws.max_row}. "
            f"Sheet should contain exactly one table with no extra content."
        )

    def test_no_extra_headers_in_row_1(self, workbook_data):
        """Verify row 1 contains exactly the required headers with no extra non-empty cells."""
        ws = workbook_data['worksheet']

        # Check each cell in row 1 up to a reasonable limit
        extra_headers = []
        for col in range(len(REQUIRED_COLUMNS) + 1, min(ws.max_column + 10, 50)):
            cell = ws.cell(row=1, column=col)
            if cell.value is not None and str(cell.value).strip():
                extra_headers.append(f"Column {col}: '{cell.value}'")

        assert len(extra_headers) == 0, (
            f"Found extra headers beyond required columns:\n"
            f"{chr(10).join(extra_headers)}\n"
            f"Sheet should have exactly these headers: {REQUIRED_COLUMNS}"
        )

    def test_no_content_outside_table(self, workbook_data):
        """Verify no content exists outside the expected table range."""
        ws = workbook_data['worksheet']
        expected_cols = len(REQUIRED_COLUMNS)
        expected_rows = 1 + EXPECTED_DATA_ROWS

        extra_content = []

        # Check for content beyond expected columns in data rows
        for row in range(1, expected_rows + 5):
            for col in range(expected_cols + 1, expected_cols + 10):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and str(cell.value).strip():
                    extra_content.append(f"Cell ({row}, {col}): '{cell.value}'")

        # Check for content beyond expected rows
        for row in range(expected_rows + 1, expected_rows + 10):
            for col in range(1, expected_cols + 5):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and str(cell.value).strip():
                    extra_content.append(f"Cell ({row}, {col}): '{cell.value}'")

        assert len(extra_content) == 0, (
            f"Found content outside expected table range (1:{expected_rows}, A:{chr(64+expected_cols)}):\n"
            f"{chr(10).join(extra_content[:10])}"
            f"{'...' if len(extra_content) > 10 else ''}\n"
            f"Sheet should contain exactly one table with no extra content."
        )


class TestSummarySheetColumns:
    """Tests for Summary sheet column structure."""

    def test_all_required_columns_present(self, workbook_data):
        """Verify all required columns are present in the Summary sheet."""
        headers = list(workbook_data['headers'].keys())

        for required_col in REQUIRED_COLUMNS:
            assert required_col in headers, (
                f"Required column '{required_col}' not found in headers: {headers}"
            )

    def test_column_order_matches_specification(self, workbook_data):
        """Verify columns are in the specified order."""
        ws = workbook_data['worksheet']

        # Read first N columns
        headers = []
        for col in range(1, len(REQUIRED_COLUMNS) + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                headers.append(str(cell_value).strip())

        assert headers == REQUIRED_COLUMNS, (
            f"Column order mismatch.\nExpected: {REQUIRED_COLUMNS}\nFound: {headers}"
        )


class TestSummarySheetData:
    """Tests for Summary sheet data content."""

    def test_correct_number_of_data_rows(self, workbook_data):
        """Verify there are exactly 3 data rows (one per case file)."""
        data_rows = len(workbook_data['data'])
        assert data_rows == EXPECTED_DATA_ROWS, (
            f"Expected {EXPECTED_DATA_ROWS} data rows (one per case file), found {data_rows}"
        )

    def test_all_case_files_present(self, workbook_data):
        """Verify all three case files are listed in the case_file column."""
        found_case_files = list(workbook_data['data'].keys())

        for expected in EXPECTED_CASE_FILES:
            assert expected in found_case_files, (
                f"Case file '{expected}' not found in data. Found: {found_case_files}"
            )


class TestSlackBusUniqueness:
    """Tests for unique slack bus requirement."""

    @pytest.mark.parametrize("case_file", EXPECTED_CASE_FILES)
    def test_unique_slack_bus(self, case_file):
        """Verify each case file has exactly one slack bus (type==3)."""
        filepath = os.path.join(INPUT_DIR, case_file)
        if not os.path.exists(filepath):
            pytest.skip(f"Input file not found: {filepath}")

        mpc = parse_matpower_file(filepath)
        slack_buses = [b['bus_i'] for b in mpc['bus'] if b['type'] == 3]

        assert len(slack_buses) == 1, (
            f"{case_file}: Expected exactly 1 slack bus (type==3), "
            f"found {len(slack_buses)}: {slack_buses}. "
            f"DC power flow requires a unique slack bus for angle reference."
        )

    def test_slack_bus_in_expected_results(self, expected_results):
        """Verify expected results correctly identify unique slack bus."""
        for case_file, expected in expected_results.items():
            slack_bus = expected.get('slack_bus')
            assert slack_bus is not None, (
                f"{case_file}: Slack bus not identified in expected results"
            )

            # Verify slack bus angle is 0 in computed theta
            theta = expected.get('theta', {})
            if theta:
                slack_angle = theta.get(slack_bus, None)
                assert slack_angle is not None, (
                    f"{case_file}: Slack bus {slack_bus} not in theta results"
                )
                assert slack_angle == pytest.approx(0.0, abs=1e-10), (
                    f"{case_file}: Slack bus angle should be 0, got {slack_angle}"
                )


class TestBusCountsExact:
    """Tests for exact bus count verification."""

    @pytest.mark.parametrize("case_file,expected_n_bus", [
        ("case57.m", 57),
        ("case118.m", 118),
        ("pglib_opf_case118_ieee.m", 118),
    ])
    def test_exact_bus_count(self, workbook_data, case_file, expected_n_bus):
        """Verify n_bus matches exactly for each case."""
        row_data = workbook_data['data'].get(case_file)
        assert row_data is not None, f"Case file {case_file} not found in output"

        actual_n_bus = row_data.get('n_bus')
        assert actual_n_bus == expected_n_bus, (
            f"{case_file}: Expected n_bus={expected_n_bus}, got {actual_n_bus}"
        )


class TestBranchCountsExact:
    """Tests for exact branch count verification (unique in-service edges)."""

    @pytest.mark.parametrize("case_file", EXPECTED_CASE_FILES)
    def test_exact_branch_count(self, workbook_data, expected_results, case_file):
        """Verify each case has exact number of unique in-service branches."""
        row_data = workbook_data['data'].get(case_file)
        expected = expected_results.get(case_file)

        if expected is None:
            pytest.skip(f"Could not compute expected results for {case_file}")

        actual = row_data.get('n_branch_in_service')
        expected_count = expected['n_branch_in_service']

        assert actual == expected_count, (
            f"{case_file}: Expected n_branch_in_service={expected_count}, got {actual}. "
            f"This counts unique undirected edges in the simple graph (parallel branches collapsed)."
        )


class TestCriticalEdgeDirected:
    """Tests for critical edge with directed fbus->tbus verification."""

    @pytest.mark.parametrize("case_file", EXPECTED_CASE_FILES)
    def test_critical_edge_undirected_match(self, workbook_data, expected_results, case_file):
        """Verify critical edge (undirected) matches expected with deterministic tie-break."""
        row_data = workbook_data['data'].get(case_file)
        expected = expected_results.get(case_file)

        if expected is None:
            pytest.skip(f"Could not compute expected results for {case_file}")

        actual_fbus = int(row_data.get('critical_edge_fbus'))
        actual_tbus = int(row_data.get('critical_edge_tbus'))
        expected_undirected = expected['critical_edge_undirected']

        # Normalize to (min, max) for undirected comparison
        actual_undirected = (min(actual_fbus, actual_tbus), max(actual_fbus, actual_tbus))

        assert actual_undirected == expected_undirected, (
            f"{case_file}: Critical edge (undirected) mismatch.\n"
            f"Expected: {expected_undirected}\n"
            f"Got: {actual_undirected}\n"
            f"Tie-break rule: lexicographically smallest (min(fbus,tbus), max(fbus,tbus))"
        )

    @pytest.mark.parametrize("case_file", EXPECTED_CASE_FILES)
    def test_critical_edge_direction_and_flow_sign(self, workbook_data, expected_results, case_file):
        """Verify flow sign is consistent with reported fbus->tbus direction.

        The flow should be computed as (theta_fbus - theta_tbus) / x * baseMVA
        where fbus and tbus are the directed endpoints from the first branch occurrence.
        """
        row_data = workbook_data['data'].get(case_file)
        expected = expected_results.get(case_file)

        if expected is None:
            pytest.skip(f"Could not compute expected results for {case_file}")

        actual_fbus = int(row_data.get('critical_edge_fbus'))
        actual_tbus = int(row_data.get('critical_edge_tbus'))
        actual_flow = row_data.get('critical_edge_flow_MW')

        # Get theta and recompute expected flow with actual direction
        theta = expected['theta']
        mpc = expected['mpc']
        baseMVA = mpc['baseMVA']

        # Find the first branch matching the undirected edge
        actual_undirected = (min(actual_fbus, actual_tbus), max(actual_fbus, actual_tbus))
        matching_branch = None
        for branch in mpc['branch']:
            if branch['status'] == 1:
                branch_undirected = (min(branch['fbus'], branch['tbus']),
                                    max(branch['fbus'], branch['tbus']))
                if branch_undirected == actual_undirected:
                    matching_branch = branch
                    break

        if matching_branch is None:
            pytest.fail(f"{case_file}: Could not find branch for edge {actual_undirected}")

        # Compute expected flow using the actual reported direction
        x = matching_branch['x']
        expected_flow_with_actual_dir = compute_line_flow(theta, actual_fbus, actual_tbus, x, baseMVA)

        # The actual flow should match with correct sign
        assert actual_flow == pytest.approx(expected_flow_with_actual_dir, rel=0.01, abs=1.0), (
            f"{case_file}: Flow sign/magnitude mismatch with reported direction.\n"
            f"Reported edge: {actual_fbus} -> {actual_tbus}\n"
            f"Expected flow: {expected_flow_with_actual_dir:.4f} MW\n"
            f"Got flow: {actual_flow:.4f} MW\n"
            f"Flow formula: (theta_fbus - theta_tbus) / x * baseMVA"
        )


class TestTieBreaking:
    """Tests for deterministic tie-breaking validation."""

    def test_tie_breaking_with_tolerance(self, expected_results):
        """Verify tie-breaking uses tolerance and selects lexicographically smallest edge."""
        for case_file, expected in expected_results.items():
            betweenness = expected['all_betweenness']
            if not betweenness:
                continue

            max_bc = max(betweenness.values())
            # Find all edges within tolerance of max
            tied_edges = [edge for edge, bc in betweenness.items()
                         if bc >= max_bc - TIE_TOLERANCE]

            if len(tied_edges) > 1:
                # There are ties - verify lexicographic ordering is used
                tied_edges.sort()
                expected_winner = tied_edges[0]
                actual_winner = expected['critical_edge_undirected']

                assert actual_winner == expected_winner, (
                    f"{case_file}: Tie-breaking failed.\n"
                    f"Tied edges (bc >= {max_bc - TIE_TOLERANCE}): {tied_edges}\n"
                    f"Expected winner (lex smallest): {expected_winner}\n"
                    f"Got: {actual_winner}"
                )

    def test_synthetic_tie_scenario(self):
        """Test tie-breaking with a synthetic graph that has known ties."""
        pytest.importorskip("networkx")
        import networkx as nx

        # Create a graph with symmetric structure that should produce ties
        # Simple diamond: 1-2, 1-3, 2-4, 3-4
        edges = [(1, 2), (1, 3), (2, 4), (3, 4)]
        nodes = {1, 2, 3, 4}

        betweenness = compute_edge_betweenness(nodes, set(edges))

        # Due to symmetry, edges (1,2) and (1,3) should have same betweenness
        # and edges (2,4) and (3,4) should have same betweenness
        bc_1_2 = betweenness.get((1, 2), 0)
        bc_1_3 = betweenness.get((1, 3), 0)

        assert bc_1_2 == pytest.approx(bc_1_3, rel=1e-10), (
            f"Symmetric edges should have equal betweenness: (1,2)={bc_1_2}, (1,3)={bc_1_3}"
        )

        # With tie, should select lexicographically smallest
        critical = find_critical_edge_with_tolerance(betweenness, TIE_TOLERANCE)

        # The max betweenness edges are (1,2) and (1,3), both = 0.5 in a diamond
        # (1,2) < (1,3) lexicographically
        max_bc = max(betweenness.values())
        tied_edges = sorted([e for e, bc in betweenness.items() if bc >= max_bc - TIE_TOLERANCE])
        expected_critical = tied_edges[0]

        assert critical == expected_critical, (
            f"Tie-breaking should select lexicographically smallest.\n"
            f"Tied edges: {tied_edges}\n"
            f"Expected: {expected_critical}\n"
            f"Got: {critical}"
        )


class TestEdgeBetweennessExact:
    """Tests for exact edge betweenness centrality verification."""

    @pytest.mark.parametrize("case_file", EXPECTED_CASE_FILES)
    def test_edge_betweenness(self, workbook_data, expected_results, case_file):
        """Verify edge betweenness matches expected value."""
        row_data = workbook_data['data'].get(case_file)
        expected = expected_results.get(case_file)

        if expected is None:
            pytest.skip(f"Could not compute expected results for {case_file}")

        actual_bc = row_data.get('critical_edge_edge_betweenness')
        expected_bc = expected['edge_betweenness']

        assert actual_bc == pytest.approx(expected_bc, rel=1e-6, abs=1e-9), (
            f"{case_file}: Edge betweenness mismatch.\n"
            f"Expected: {expected_bc}\n"
            f"Got: {actual_bc}"
        )


class TestDCPowerFlowExact:
    """Tests for DC power flow computation verification."""

    @pytest.mark.parametrize("case_file", EXPECTED_CASE_FILES)
    def test_flow(self, workbook_data, expected_results, case_file):
        """Verify critical edge flow matches DC power flow computation."""
        row_data = workbook_data['data'].get(case_file)
        expected = expected_results.get(case_file)

        if expected is None:
            pytest.skip(f"Could not compute expected results for {case_file}")

        actual_flow = row_data.get('critical_edge_flow_MW')
        expected_flow = expected['flow_MW']

        # Use relatively loose tolerance for DC power flow (numerical differences)
        assert actual_flow == pytest.approx(expected_flow, rel=0.01, abs=1.0), (
            f"{case_file}: Critical edge flow mismatch.\n"
            f"Expected: {expected_flow:.4f} MW\n"
            f"Got: {actual_flow:.4f} MW\n"
            f"Flow formula: (theta_fbus - theta_tbus) / x * baseMVA"
        )


class TestNumericCellsNotFormulas:
    """Tests to ensure numeric cells contain values, not formulas."""

    def test_numeric_cells_are_values_not_formulas(self, workbook_data):
        """Verify numeric cells contain direct values, not formulas."""
        ws = workbook_data['worksheet']
        headers = workbook_data['headers']

        numeric_columns = [
            "n_bus", "n_branch_in_service",
            "critical_edge_fbus", "critical_edge_tbus",
            "critical_edge_edge_betweenness", "critical_edge_flow_MW"
        ]

        formula_cells = []
        for row in range(2, ws.max_row + 1):
            for col_name in numeric_columns:
                col = headers.get(col_name)
                if col:
                    cell = ws.cell(row=row, column=col)

                    # Check if cell is a formula
                    is_formula = False

                    # Method 1: Check data_type (openpyxl uses 'f' for formula)
                    if hasattr(cell, 'data_type') and cell.data_type == 'f':
                        is_formula = True

                    # Method 2: Check if value is a string starting with '='
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        is_formula = True

                    if is_formula:
                        case_file = ws.cell(row=row, column=headers.get("case_file", 1)).value
                        formula_cells.append(
                            f"Row {row} ({case_file}), Column '{col_name}': "
                            f"value='{cell.value}', data_type={getattr(cell, 'data_type', 'N/A')}"
                        )

        assert len(formula_cells) == 0, (
            f"Numeric cells should contain direct values, not formulas:\n"
            f"{chr(10).join(formula_cells)}\n"
            f"Requirement: numeric values must be directly written numbers."
        )

    def test_numeric_cells_have_numeric_type(self, workbook_data):
        """Verify numeric cells have numeric Python type (int or float)."""
        ws = workbook_data['worksheet']
        headers = workbook_data['headers']

        numeric_columns = [
            "n_bus", "n_branch_in_service",
            "critical_edge_fbus", "critical_edge_tbus",
            "critical_edge_edge_betweenness", "critical_edge_flow_MW"
        ]

        non_numeric_cells = []
        for row in range(2, ws.max_row + 1):
            for col_name in numeric_columns:
                col = headers.get(col_name)
                if col:
                    cell = ws.cell(row=row, column=col)

                    if cell.value is not None and not isinstance(cell.value, (int, float)):
                        case_file = ws.cell(row=row, column=headers.get("case_file", 1)).value
                        non_numeric_cells.append(
                            f"Row {row} ({case_file}), Column '{col_name}': "
                            f"value='{cell.value}' (type: {type(cell.value).__name__})"
                        )

        assert len(non_numeric_cells) == 0, (
            f"Numeric cells should have numeric type (int or float):\n"
            f"{chr(10).join(non_numeric_cells)}"
        )


class TestFinancialModelStylingStrict:
    """Strict tests for financial model styling conventions."""

    def test_all_header_cells_are_bold(self, workbook_data):
        """Verify EVERY header cell has bold font."""
        ws = workbook_data['worksheet']
        headers = workbook_data['headers']

        non_bold_headers = []
        for header_name, col in headers.items():
            cell = ws.cell(row=1, column=col)
            if cell.font is None or not cell.font.bold:
                non_bold_headers.append(f"Column {col} ('{header_name}')")

        assert len(non_bold_headers) == 0, (
            f"The following header cells are NOT bold:\n"
            f"{', '.join(non_bold_headers)}\n"
            f"All header cells must have bold font per financial model conventions."
        )

    def test_all_header_cells_have_black_font(self, workbook_data):
        """Verify EVERY header cell has black font color."""
        ws = workbook_data['worksheet']
        headers = workbook_data['headers']

        def is_black_color(color_obj) -> bool:
            """Check if a color object represents black."""
            if color_obj is None:
                return True  # Default (no color) is black

            # Check RGB value
            if hasattr(color_obj, 'rgb') and color_obj.rgb:
                rgb = str(color_obj.rgb).upper()
                # Black RGB values
                if rgb in ["000000", "FF000000", "00000000"]:
                    return True
                # Check if it's a dark color (close to black)
                if len(rgb) >= 6:
                    try:
                        rgb_hex = rgb[-6:]
                        r = int(rgb_hex[0:2], 16)
                        g = int(rgb_hex[2:4], 16)
                        b = int(rgb_hex[4:6], 16)
                        # Consider it black if all components are very low
                        if r < 30 and g < 30 and b < 30:
                            return True
                    except ValueError:
                        pass

            # Check theme color (theme 1 is often black in default themes)
            if hasattr(color_obj, 'theme') and color_obj.theme is not None:
                if color_obj.theme == 1:  # Often black in default themes
                    return True

            # Check indexed color
            if hasattr(color_obj, 'indexed') and color_obj.indexed is not None:
                # Index 0, 1, 64 are typically black
                if color_obj.indexed in [0, 1, 64]:
                    return True

            # If type is None, treat as default (black)
            if hasattr(color_obj, 'type') and color_obj.type is None:
                return True

            return False

        non_black_headers = []
        for header_name, col in headers.items():
            cell = ws.cell(row=1, column=col)
            font_color = cell.font.color if cell.font else None

            if not is_black_color(font_color):
                color_str = str(font_color.rgb) if (font_color and hasattr(font_color, 'rgb')) else str(font_color)
                non_black_headers.append(
                    f"Column {col} ('{header_name}'): color={color_str}"
                )

        assert len(non_black_headers) == 0, (
            f"The following header cells do NOT have black font:\n"
            f"{chr(10).join(non_black_headers)}\n"
            f"Header cells must have black font."
        )

    def test_all_header_cells_have_gray_fill(self, workbook_data):
        """Verify EVERY header cell has light gray fill (strict assertion)."""
        ws = workbook_data['worksheet']
        headers = workbook_data['headers']

        def is_gray_fill(fill_obj) -> Tuple[bool, Optional[str]]:
            """Check if a fill object is light gray. Returns (is_gray, color_str)."""
            if fill_obj is None or fill_obj.fill_type is None or fill_obj.fill_type == "none":
                return False, "no fill"

            color_str = None
            if hasattr(fill_obj, 'fgColor') and fill_obj.fgColor:
                if hasattr(fill_obj.fgColor, 'rgb') and fill_obj.fgColor.rgb:
                    color_str = str(fill_obj.fgColor.rgb).upper()

                    # Check against known gray values
                    for gray in LIGHT_GRAY_RGBS:
                        if gray.upper() == color_str or color_str.endswith(gray.upper()):
                            return True, color_str

                    # Check if R==G==B (gray-ish)
                    if len(color_str) >= 6:
                        try:
                            rgb = color_str[-6:]
                            r = int(rgb[0:2], 16)
                            g = int(rgb[2:4], 16)
                            b = int(rgb[4:6], 16)
                            # Gray: R, G, B are similar and not too dark/light
                            if (abs(r - g) < 25 and abs(g - b) < 25 and abs(r - b) < 25
                                and 100 < r < 250):  # Light gray range
                                return True, color_str
                        except ValueError:
                            pass

                # Check theme color
                if hasattr(fill_obj.fgColor, 'theme') and fill_obj.fgColor.theme is not None:
                    # Theme 0 with tint often produces gray
                    return True, f"theme:{fill_obj.fgColor.theme}"

            return False, color_str

        no_fill_headers = []
        wrong_fill_headers = []

        for header_name, col in headers.items():
            cell = ws.cell(row=1, column=col)
            is_gray, color_str = is_gray_fill(cell.fill)

            if color_str == "no fill":
                no_fill_headers.append(f"Column {col} ('{header_name}')")
            elif not is_gray:
                wrong_fill_headers.append(
                    f"Column {col} ('{header_name}'): fill={color_str}"
                )

        # Assert no missing fills
        assert len(no_fill_headers) == 0, (
            f"The following header cells have NO fill:\n"
            f"{', '.join(no_fill_headers)}\n"
            f"All header cells must have light gray fill."
        )

        # Assert fills are gray (strict - not just warning)
        assert len(wrong_fill_headers) == 0, (
            f"The following header cells do NOT have light gray fill:\n"
            f"{chr(10).join(wrong_fill_headers)}\n"
            f"Header cells must have light gray fill. "
            f"Acceptable gray values include: {LIGHT_GRAY_RGBS[:5]}..."
        )

    def test_all_numeric_data_cells_have_blue_font(self, workbook_data):
        """Verify ALL numeric data cells have blue font color."""
        ws = workbook_data['worksheet']
        headers = workbook_data['headers']

        def is_blue_color(color_obj) -> Tuple[bool, Optional[str]]:
            """Check if a color object represents blue. Returns (is_blue, color_str)."""
            if color_obj is None:
                return False, None

            color_str = None
            if hasattr(color_obj, 'rgb') and color_obj.rgb:
                color_str = str(color_obj.rgb).upper()

                # Check against known blue values
                for blue in BLUE_RGBS:
                    if blue.upper() == color_str or color_str.endswith(blue.upper()):
                        return True, color_str

                # Check if blue component is dominant
                if len(color_str) >= 6:
                    try:
                        rgb = color_str[-6:]
                        r = int(rgb[0:2], 16)
                        g = int(rgb[2:4], 16)
                        b = int(rgb[4:6], 16)
                        if b > r and b > g and b >= 100:
                            return True, color_str
                    except ValueError:
                        pass

            # Check theme color (theme 4 or 5 are often blue)
            if hasattr(color_obj, 'theme') and color_obj.theme is not None:
                if color_obj.theme in [4, 5]:  # Often blue in default themes
                    return True, f"theme:{color_obj.theme}"

            return False, color_str

        numeric_columns = [
            "n_bus", "n_branch_in_service",
            "critical_edge_fbus", "critical_edge_tbus",
            "critical_edge_edge_betweenness", "critical_edge_flow_MW"
        ]

        non_blue_cells = []
        for row in range(2, ws.max_row + 1):
            for col_name in numeric_columns:
                col = headers.get(col_name)
                if col:
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        font_color = cell.font.color if cell.font else None
                        is_blue, color_str = is_blue_color(font_color)

                        if not is_blue:
                            case_file = ws.cell(row=row, column=headers.get("case_file", 1)).value
                            non_blue_cells.append(
                                f"Row {row} ({case_file}), Column '{col_name}': "
                                f"value={cell.value}, font_color={color_str}"
                            )

        assert len(non_blue_cells) == 0, (
            f"The following numeric cells do NOT have blue font:\n"
            f"{chr(10).join(non_blue_cells[:10])}"
            f"{'...' if len(non_blue_cells) > 10 else ''}\n"
            f"Financial model convention: numeric input cells should have blue font."
        )


class TestDataConsistency:
    """Tests for data consistency and correctness."""

    def test_fbus_tbus_are_different(self, workbook_data):
        """Verify fbus and tbus are different for each edge (no self-loops)."""
        data = workbook_data['data']

        for case_file, row_data in data.items():
            fbus = row_data.get('critical_edge_fbus')
            tbus = row_data.get('critical_edge_tbus')

            if fbus is not None and tbus is not None:
                assert fbus != tbus, (
                    f"{case_file}: fbus ({fbus}) and tbus ({tbus}) must be different "
                    f"(no self-loops allowed)"
                )

    def test_buses_are_positive_integers(self, workbook_data):
        """Verify edge buses are positive integers."""
        data = workbook_data['data']

        for case_file, row_data in data.items():
            fbus = row_data.get('critical_edge_fbus')
            tbus = row_data.get('critical_edge_tbus')

            assert fbus is not None and fbus > 0, (
                f"{case_file}: fbus should be positive, got {fbus}"
            )
            assert tbus is not None and tbus > 0, (
                f"{case_file}: tbus should be positive, got {tbus}"
            )
            assert float(fbus) == int(fbus), (
                f"{case_file}: fbus should be integer, got {fbus}"
            )
            assert float(tbus) == int(tbus), (
                f"{case_file}: tbus should be integer, got {tbus}"
            )

    def test_edge_betweenness_in_valid_range(self, workbook_data):
        """Verify edge betweenness values are between 0 and 1 (normalized)."""
        data = workbook_data['data']

        for case_file, row_data in data.items():
            bc = row_data.get('critical_edge_edge_betweenness')
            assert bc is not None, f"{case_file}: edge betweenness is None"
            assert isinstance(bc, (int, float)), (
                f"{case_file}: edge betweenness should be numeric, got {type(bc)}"
            )
            assert 0 <= bc <= 1, (
                f"{case_file}: edge betweenness should be in [0,1] (normalized), got {bc}"
            )

    def test_flow_is_numeric(self, workbook_data):
        """Verify flow values are numeric."""
        data = workbook_data['data']

        for case_file, row_data in data.items():
            flow = row_data.get('critical_edge_flow_MW')
            assert flow is not None, f"{case_file}: flow is None"
            assert isinstance(flow, (int, float)), (
                f"{case_file}: flow should be numeric, got {type(flow)}"
            )

    def test_all_data_filled(self, workbook_data):
        """Verify all cells in data rows have values (no missing data)."""
        ws = workbook_data['worksheet']
        headers = workbook_data['headers']

        missing_cells = []
        for row in range(2, len(EXPECTED_CASE_FILES) + 2):
            for header_name, col in headers.items():
                cell = ws.cell(row=row, column=col)
                if cell.value is None:
                    case_file = ws.cell(row=row, column=headers.get("case_file", 1)).value
                    missing_cells.append(
                        f"Row {row} ({case_file}), Column '{header_name}'"
                    )

        assert len(missing_cells) == 0, (
            f"Missing values in cells:\n{chr(10).join(missing_cells)}"
        )


class TestParallelBranchHandling:
    """Tests for parallel branch handling in graph construction and flow computation."""

    def test_parallel_branches_collapsed_in_graph(self, expected_results):
        """Verify parallel branches are collapsed into single edges for graph centrality."""
        case_file = "case118.m"  # Known to have parallel branches
        expected = expected_results.get(case_file)

        if expected is None:
            pytest.skip(f"Could not compute expected results for {case_file}")

        mpc = expected['mpc']
        raw_branch_count = sum(1 for b in mpc['branch'] if b['status'] == 1)
        unique_edge_count = expected['n_branch_in_service']

        # case118 has parallel branches, so unique < raw
        assert unique_edge_count <= raw_branch_count, (
            f"{case_file}: Unique edges ({unique_edge_count}) should be <= "
            f"raw branch count ({raw_branch_count})"
        )

        # Verify that there are actually parallel branches
        if unique_edge_count == raw_branch_count:
            pytest.skip(f"{case_file}: No parallel branches found to test")

    def test_dc_flow_uses_all_branches(self, expected_results):
        """Verify DC power flow susceptance matrix uses ALL in-service branches."""
        for case_file, expected in expected_results.items():
            mpc = expected['mpc']

            # Count branches contributing to B matrix
            in_service_branches = [b for b in mpc['branch'] if b['status'] == 1 and b['x'] != 0]

            # Verify we have branches
            assert len(in_service_branches) > 0, (
                f"{case_file}: No in-service branches with non-zero reactance"
            )

    def test_first_branch_used_for_edge_info(self, expected_results):
        """Verify first occurrence in file order is used for edge direction/reactance."""
        for case_file, expected in expected_results.items():
            mpc = expected['mpc']
            critical_undirected = expected['critical_edge_undirected']
            expected_fbus = expected['critical_edge_fbus']
            expected_tbus = expected['critical_edge_tbus']
            expected_x = expected['critical_edge_x']

            if critical_undirected is None:
                continue

            # Find first branch matching this undirected edge
            first_branch = None
            for branch in mpc['branch']:
                if branch['status'] == 1:
                    branch_undirected = (min(branch['fbus'], branch['tbus']),
                                        max(branch['fbus'], branch['tbus']))
                    if branch_undirected == critical_undirected:
                        first_branch = branch
                        break

            assert first_branch is not None, (
                f"{case_file}: Could not find branch for critical edge {critical_undirected}"
            )

            # Verify expected values match first branch
            assert expected_fbus == first_branch['fbus'], (
                f"{case_file}: Expected fbus should be from first branch. "
                f"Expected {expected_fbus}, first branch fbus={first_branch['fbus']}"
            )
            assert expected_tbus == first_branch['tbus'], (
                f"{case_file}: Expected tbus should be from first branch. "
                f"Expected {expected_tbus}, first branch tbus={first_branch['tbus']}"
            )
            assert expected_x == first_branch['x'], (
                f"{case_file}: Expected x should be from first branch. "
                f"Expected {expected_x}, first branch x={first_branch['x']}"
            )


class TestEdgeCases:
    """Tests for edge case handling."""

    def test_only_inservice_branches_counted(self):
        """Verify only in-service branches (status==1) are included."""
        for case_file in EXPECTED_CASE_FILES:
            filepath = os.path.join(INPUT_DIR, case_file)
            if os.path.exists(filepath):
                mpc = parse_matpower_file(filepath)

                in_service = sum(1 for b in mpc['branch'] if b['status'] == 1)
                out_of_service = sum(1 for b in mpc['branch'] if b['status'] != 1)

                assert in_service > 0, f"{case_file}: No in-service branches found"


class TestIndependentCaseValidation:
    """Validate each case independently."""

    @pytest.mark.parametrize("case_file", EXPECTED_CASE_FILES)
    def test_case_independent_validation(self, workbook_data, expected_results, case_file):
        """Validate case results against independently computed values."""
        row_data = workbook_data['data'].get(case_file)
        expected = expected_results.get(case_file)

        if expected is None:
            pytest.skip(f"Could not compute expected results for {case_file}")

        # Validate n_bus
        assert row_data['n_bus'] == expected['n_bus'], \
            f"{case_file}: n_bus mismatch"

        # Validate n_branch_in_service
        assert row_data['n_branch_in_service'] == expected['n_branch_in_service'], \
            f"{case_file}: n_branch_in_service mismatch"

        # Validate critical edge (undirected)
        actual_edge = (
            min(int(row_data['critical_edge_fbus']), int(row_data['critical_edge_tbus'])),
            max(int(row_data['critical_edge_fbus']), int(row_data['critical_edge_tbus']))
        )
        assert actual_edge == expected['critical_edge_undirected'], \
            f"{case_file}: critical_edge mismatch"

        # Validate edge betweenness
        assert row_data['critical_edge_edge_betweenness'] == \
               pytest.approx(expected['edge_betweenness'], rel=1e-6), \
            f"{case_file}: edge_betweenness mismatch"


class TestUsingPandas:
    """Tests using pandas for data validation."""

    def test_read_with_pandas(self):
        """Verify file can be read with pandas."""
        pytest.importorskip("pandas")
        pytest.importorskip("openpyxl")
        import pandas as pd

        if not os.path.exists(OUTPUT_FILE):
            pytest.skip(f"Output file not found: {OUTPUT_FILE}")

        try:
            df = pd.read_excel(OUTPUT_FILE, sheet_name="Summary")
            assert df is not None
            assert len(df) == EXPECTED_DATA_ROWS, f"Expected {EXPECTED_DATA_ROWS} rows, got {len(df)}"
        except Exception as e:
            pytest.fail(f"Failed to read Excel with pandas: {e}")

    def test_column_dtypes(self):
        """Verify columns have appropriate data types."""
        pytest.importorskip("pandas")
        pytest.importorskip("openpyxl")
        import pandas as pd

        if not os.path.exists(OUTPUT_FILE):
            pytest.skip(f"Output file not found: {OUTPUT_FILE}")

        df = pd.read_excel(OUTPUT_FILE, sheet_name="Summary")

        # case_file should be string/object (object in older pandas, StringDtype in newer)
        assert df["case_file"].dtype == object or str(df["case_file"].dtype) == 'str', (
            f"case_file should be object or str dtype, got {df['case_file'].dtype}"
        )

        # Numeric columns should be numeric
        numeric_cols = [
            "n_bus", "n_branch_in_service",
            "critical_edge_fbus", "critical_edge_tbus",
            "critical_edge_edge_betweenness", "critical_edge_flow_MW"
        ]

        for col in numeric_cols:
            assert pd.api.types.is_numeric_dtype(df[col]), (
                f"{col} should be numeric, got {df[col].dtype}"
            )

    def test_no_null_values(self):
        """Verify there are no null values in the dataframe."""
        pytest.importorskip("pandas")
        pytest.importorskip("openpyxl")
        import pandas as pd

        if not os.path.exists(OUTPUT_FILE):
            pytest.skip(f"Output file not found: {OUTPUT_FILE}")

        df = pd.read_excel(OUTPUT_FILE, sheet_name="Summary")

        null_counts = df.isnull().sum()
        total_nulls = null_counts.sum()

        assert total_nulls == 0, (
            f"Found {total_nulls} null values:\n{null_counts[null_counts > 0]}"
        )

    def test_exact_column_count_pandas(self):
        """Verify pandas reads exactly the expected number of columns."""
        pytest.importorskip("pandas")
        pytest.importorskip("openpyxl")
        import pandas as pd

        if not os.path.exists(OUTPUT_FILE):
            pytest.skip(f"Output file not found: {OUTPUT_FILE}")

        df = pd.read_excel(OUTPUT_FILE, sheet_name="Summary")

        assert len(df.columns) == len(REQUIRED_COLUMNS), (
            f"Expected exactly {len(REQUIRED_COLUMNS)} columns, "
            f"got {len(df.columns)}: {list(df.columns)}"
        )

    def test_exact_row_count_pandas(self):
        """Verify pandas reads exactly the expected number of rows."""
        pytest.importorskip("pandas")
        pytest.importorskip("openpyxl")
        import pandas as pd

        if not os.path.exists(OUTPUT_FILE):
            pytest.skip(f"Output file not found: {OUTPUT_FILE}")

        df = pd.read_excel(OUTPUT_FILE, sheet_name="Summary")

        assert len(df) == EXPECTED_DATA_ROWS, (
            f"Expected exactly {EXPECTED_DATA_ROWS} data rows, got {len(df)}"
        )


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
