"""Stronger expectation tests for the FTP bruteforce comparison task.

This version recomputes the expected values directly from:
- /root/bruteforce.pcap
- /root/ftp.pcap
- /root/samplecaptures  (Wireshark SampleCaptures HTML)

It validates:
- exact workbook structure
- exact comparison-sheet headers and row coverage
- exact per-PCAP computed metrics
- exact protocol category extracted from the HTML page
- methods sheet constraints
- required Excel formatting rules
"""

from __future__ import annotations

import json
import math
import os
import re
from collections import Counter, defaultdict
from pathlib import Path

import pytest
from openpyxl import load_workbook

try:
    from scapy.all import IP, TCP, Raw, PcapReader  # type: ignore
    HAVE_SCAPY = True
except Exception:
    HAVE_SCAPY = False

try:
    import dpkt  # type: ignore
    HAVE_DPKT = True
except Exception:
    HAVE_DPKT = False


OUTPUT_FILE = Path("/root/ftp_bruteforce_comparison.xlsx")
BRUTEFORCE_PCAP = Path("/root/bruteforce.pcap")
FTP_PCAP = Path("/root/ftp.pcap")
SAMPLECAPTURES_PATH = Path("/root/samplecaptures")

EXPECTED_SHEETS = ["comparison", "methods"]
EXPECTED_COLUMNS = [
    "pcap",
    "protocol_category",
    "server_ip",
    "client_ip",
    "total_login_attempts",
    "unique_usernames",
    "most_targeted_username",
    "top_password_candidate",
    "success_count",
    "bruteforce_score",
]
EXPECTED_PCAPS = ["bruteforce.pcap", "ftp.pcap"]
NUMERIC_COLUMNS = [
    "total_login_attempts",
    "unique_usernames",
    "success_count",
    "bruteforce_score",
]


def _stream_id(src: str, sport: int, dst: str, dport: int) -> tuple:
    """Normalize a TCP connection identity irrespective of direction."""
    a = (src, int(sport))
    b = (dst, int(dport))
    return tuple(sorted((a, b)))


def _iter_control_packets(path: Path):
    """Yield ordered FTP control-channel payload packets from a PCAP."""
    assert path.exists(), f"Missing PCAP file: {path}"

    if HAVE_SCAPY:
        with PcapReader(str(path)) as pcap:
            for pkt in pcap:
                if IP not in pkt or TCP not in pkt:
                    continue
                ip = pkt[IP]
                tcp = pkt[TCP]
                if int(tcp.sport) != 21 and int(tcp.dport) != 21:
                    continue
                payload = b""
                if Raw in pkt:
                    payload = bytes(pkt[Raw].load)
                else:
                    try:
                        payload = bytes(tcp.payload)
                    except Exception:
                        payload = b""
                if not payload:
                    continue
                yield {
                    "src": str(ip.src),
                    "dst": str(ip.dst),
                    "sport": int(tcp.sport),
                    "dport": int(tcp.dport),
                    "payload": payload,
                }
        return

    if HAVE_DPKT:
        with path.open("rb") as f:
            pcap = dpkt.pcap.Reader(f)
            for _, buf in pcap:
                eth = dpkt.ethernet.Ethernet(buf)
                ip = getattr(eth, "data", None)
                if ip is None or not isinstance(ip, dpkt.ip.IP):
                    continue
                tcp = getattr(ip, "data", None)
                if tcp is None or not isinstance(tcp, dpkt.tcp.TCP):
                    continue
                if int(tcp.sport) != 21 and int(tcp.dport) != 21:
                    continue
                payload = bytes(tcp.data or b"")
                if not payload:
                    continue
                yield {
                    "src": ".".join(str(b) for b in ip.src),
                    "dst": ".".join(str(b) for b in ip.dst),
                    "sport": int(tcp.sport),
                    "dport": int(tcp.dport),
                    "payload": payload,
                }
        return

    pytest.fail("Neither scapy nor dpkt is available to parse PCAP files.")


def _split_lines_from_buffer(buffer: bytes):
    """Extract CRLF-terminated FTP control lines from a directional byte buffer."""
    lines = []
    while b"\n" in buffer:
        line, buffer = buffer.split(b"\n", 1)
        line = line.rstrip(b"\r")
        text = line.decode("utf-8", errors="ignore").strip("\x00")
        if text != "":
            lines.append(text)
    return lines, buffer


def analyze_ftp_pcap(path: Path) -> dict:
    """Parse FTP control traffic and compute all required metrics for one PCAP."""
    streams = {}

    for pkt in _iter_control_packets(path):
        sid = _stream_id(pkt["src"], pkt["sport"], pkt["dst"], pkt["dport"])
        stream = streams.setdefault(
            sid,
            {
                "buffers": defaultdict(bytes),
                "events": [],
            },
        )

        dkey = (pkt["src"], pkt["sport"], pkt["dst"], pkt["dport"])
        stream["buffers"][dkey] += pkt["payload"]
        lines, remainder = _split_lines_from_buffer(stream["buffers"][dkey])
        stream["buffers"][dkey] = remainder

        for line in lines:
            stream["events"].append(
                {
                    "src": pkt["src"],
                    "dst": pkt["dst"],
                    "sport": pkt["sport"],
                    "dport": pkt["dport"],
                    "line": line,
                }
            )

    qualifying_streams = []
    for stream in streams.values():
        banner_event = None
        for ev in stream["events"]:
            if re.match(r"^220(?:\s|-)", ev["line"]):
                banner_event = ev
                break
        if banner_event is None:
            continue

        stream["server_ip"] = banner_event["src"]
        stream["client_ip"] = banner_event["dst"]
        qualifying_streams.append(stream)

    assert qualifying_streams, f"No FTP control stream with a 220 banner found in {path}"

    server_ips = {stream["server_ip"] for stream in qualifying_streams}
    client_ips = {stream["client_ip"] for stream in qualifying_streams}
    assert len(server_ips) == 1, f"Expected one server IP in {path}, got {server_ips}"
    assert len(client_ips) == 1, f"Expected one client IP in {path}, got {client_ips}"

    server_ip = next(iter(server_ips))
    client_ip = next(iter(client_ips))

    user_counter = Counter()
    pass_counter = Counter()
    total_login_attempts = 0
    success_count = 0

    for stream in qualifying_streams:
        pending_pass = False
        for ev in stream["events"]:
            line = ev["line"]

            if ev["src"] == stream["client_ip"] and ev["dst"] == stream["server_ip"]:
                upper = line.upper()
                if upper.startswith("USER "):
                    username = line[5:].strip()
                    user_counter[username] += 1
                elif upper.startswith("PASS "):
                    password = line[5:].strip()
                    pass_counter[password] += 1
                    total_login_attempts += 1
                    pending_pass = True

            elif ev["src"] == stream["server_ip"] and ev["dst"] == stream["client_ip"]:
                if pending_pass and re.match(r"^\d{3}(?:\s|-)", line):
                    if line.startswith("230"):
                        success_count += 1
                    pending_pass = False

    unique_usernames = len(user_counter)
    most_targeted_username = min(
        user_counter.items(),
        key=lambda item: (-item[1], item[0]),
    )[0] if user_counter else ""

    top_password_candidate = min(
        pass_counter.items(),
        key=lambda item: (-item[1], item[0]),
    )[0] if pass_counter else ""

    bruteforce_score = round(total_login_attempts / max(1, unique_usernames), 4)

    return {
        "server_ip": server_ip,
        "client_ip": client_ip,
        "total_login_attempts": int(total_login_attempts),
        "unique_usernames": int(unique_usernames),
        "most_targeted_username": most_targeted_username,
        "top_password_candidate": top_password_candidate,
        "success_count": int(success_count),
        "bruteforce_score": bruteforce_score,
    }


def _find_samplecaptures_html() -> Path:
    """Find the HTML page inside /root/samplecaptures."""
    assert SAMPLECAPTURES_PATH.exists(), f"Missing samplecaptures path: {SAMPLECAPTURES_PATH}"

    if SAMPLECAPTURES_PATH.is_file():
        return SAMPLECAPTURES_PATH

    candidates = [
        p for p in SAMPLECAPTURES_PATH.rglob("*")
        if p.is_file() and p.suffix.lower() in {".html", ".htm"}
    ]
    assert candidates, f"No HTML file found under {SAMPLECAPTURES_PATH}"

    preferred = [
        p for p in candidates
        if p.name.lower() in {"index.html", "index.htm", "samplecaptures.html", "sample-captures.html"}
    ]
    return sorted(preferred or candidates)[0]


def extract_protocol_category() -> str:
    """Extract the visible section/category heading corresponding to FTP from the HTML page."""
    html_path = _find_samplecaptures_html()
    html = html_path.read_text(encoding="utf-8", errors="ignore")

    try:
        from bs4 import BeautifulSoup  # type: ignore

        soup = BeautifulSoup(html, "html.parser")

        # Look for text nodes or anchors containing FTP, then use the nearest previous heading.
        matches = []

        for tag in soup.find_all(True):
            text = " ".join(tag.get_text(" ", strip=True).split())
            if not text:
                continue
            if re.search(r"\bFTP\b", text, flags=re.IGNORECASE):
                heading = tag.find_previous(["h1", "h2", "h3", "h4", "h5", "h6"])
                if heading is not None:
                    heading_text = " ".join(heading.get_text(" ", strip=True).split())
                    if heading_text:
                        matches.append(heading_text)

        if matches:
            return matches[0]
    except Exception:
        pass

    # Regex fallback: scan for heading tags and remember the last visible heading.
    heading_pattern = re.compile(r"<h([1-6])[^>]*>(.*?)</h\1>", re.IGNORECASE | re.DOTALL)
    ftp_pattern = re.compile(r"\bFTP\b", re.IGNORECASE)

    headings = []
    for match in heading_pattern.finditer(html):
        heading_text = re.sub(r"<[^>]+>", " ", match.group(2))
        heading_text = " ".join(heading_text.split())
        headings.append((match.start(), heading_text))

    assert headings, "No visible heading tags found in samplecaptures HTML"

    ftp_match = ftp_pattern.search(html)
    assert ftp_match is not None, "Could not find 'FTP' in samplecaptures HTML"

    ftp_pos = ftp_match.start()
    prior_headings = [text for pos, text in headings if pos < ftp_pos and text]
    assert prior_headings, "Could not find a prior heading for the FTP section in samplecaptures HTML"

    return prior_headings[-1]


def compute_expected_rows() -> dict:
    """Recompute expected comparison rows for both PCAPs."""
    protocol_category = extract_protocol_category()

    results = {}
    for pcap_name, pcap_path in [
        ("bruteforce.pcap", BRUTEFORCE_PCAP),
        ("ftp.pcap", FTP_PCAP),
    ]:
        analysis = analyze_ftp_pcap(pcap_path)
        results[pcap_name] = {
            "pcap": pcap_name,
            "protocol_category": protocol_category,
            **analysis,
        }
    return results


@pytest.fixture(scope="module")
def expected_rows():
    return compute_expected_rows()


@pytest.fixture(scope="module")
def workbook():
    assert OUTPUT_FILE.exists(), f"Output file not found at {OUTPUT_FILE}"
    assert OUTPUT_FILE.stat().st_size > 0, "Output file is empty"
    wb = load_workbook(OUTPUT_FILE)
    yield wb
    wb.close()


@pytest.fixture(scope="module")
def comparison_sheet(workbook):
    assert workbook.sheetnames == EXPECTED_SHEETS, (
        f"Workbook sheet names/order must be exactly {EXPECTED_SHEETS}, got {workbook.sheetnames}"
    )
    return workbook["comparison"]


@pytest.fixture(scope="module")
def methods_sheet(workbook):
    return workbook["methods"]


def _comparison_rows_as_dicts(ws):
    headers = [cell.value for cell in ws[1]]
    assert headers == EXPECTED_COLUMNS, (
        f"comparison header row must be exactly {EXPECTED_COLUMNS}, got {headers}"
    )

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None and v != "" for v in row):
            row_dict = {headers[i]: row[i] for i in range(len(headers))}
            rows.append(row_dict)
    return rows


def _comparison_by_pcap(ws):
    rows = _comparison_rows_as_dicts(ws)
    assert len(rows) == 2, f"comparison sheet must have exactly 2 data rows, got {len(rows)}"
    keyed = {}
    for row in rows:
        pcap = str(row["pcap"])
        assert pcap in EXPECTED_PCAPS, f"Unexpected pcap value in comparison sheet: {pcap!r}"
        assert pcap not in keyed, f"Duplicate pcap row found in comparison sheet: {pcap}"
        keyed[pcap] = row
    assert set(keyed.keys()) == set(EXPECTED_PCAPS), (
        f"comparison sheet must contain exactly {EXPECTED_PCAPS}, got {sorted(keyed.keys())}"
    )
    return keyed


class TestOutputFile:
    def test_output_file_exists(self):
        assert OUTPUT_FILE.exists(), f"Output file not found at {OUTPUT_FILE}"

    def test_output_file_not_empty(self):
        assert OUTPUT_FILE.exists(), f"Output file not found at {OUTPUT_FILE}"
        assert OUTPUT_FILE.stat().st_size > 0, "Output file is empty"

    def test_valid_excel_workbook(self, workbook):
        assert workbook is not None


class TestWorkbookStructure:
    def test_exact_sheet_names_and_order(self, workbook):
        assert workbook.sheetnames == EXPECTED_SHEETS, (
            f"Workbook sheet names/order must be exactly {EXPECTED_SHEETS}, got {workbook.sheetnames}"
        )


class TestComparisonSheetStructure:
    def test_exact_header_row(self, comparison_sheet):
        headers = [cell.value for cell in comparison_sheet[1]]
        assert headers == EXPECTED_COLUMNS, (
            f"comparison header row must be exactly {EXPECTED_COLUMNS}, got {headers}"
        )

    def test_exact_two_data_rows(self, comparison_sheet):
        rows = _comparison_rows_as_dicts(comparison_sheet)
        assert len(rows) == 2, f"comparison sheet must have exactly 2 data rows, got {len(rows)}"

    def test_contains_both_pcaps(self, comparison_sheet):
        rows = _comparison_by_pcap(comparison_sheet)
        assert set(rows) == set(EXPECTED_PCAPS)


class TestComparisonSheetExactValues:
    def test_exact_values_per_pcap(self, comparison_sheet, expected_rows):
        actual = _comparison_by_pcap(comparison_sheet)

        for pcap_name, expected in expected_rows.items():
            row = actual[pcap_name]

            assert str(row["pcap"]) == expected["pcap"]
            assert str(row["protocol_category"]) == expected["protocol_category"]
            assert str(row["server_ip"]) == expected["server_ip"]
            assert str(row["client_ip"]) == expected["client_ip"]
            assert int(row["total_login_attempts"]) == expected["total_login_attempts"]
            assert int(row["unique_usernames"]) == expected["unique_usernames"]
            assert str(row["most_targeted_username"]) == expected["most_targeted_username"]
            assert str(row["top_password_candidate"]) == expected["top_password_candidate"]
            assert int(row["success_count"]) == expected["success_count"]
            assert float(row["bruteforce_score"]) == pytest.approx(
                expected["bruteforce_score"], rel=0, abs=1e-4
            ), (
                f"bruteforce_score mismatch for {pcap_name}: "
                f"expected {expected['bruteforce_score']}, got {row['bruteforce_score']}"
            )

    def test_bruteforce_score_formula(self, comparison_sheet):
        actual = _comparison_by_pcap(comparison_sheet)
        for pcap_name, row in actual.items():
            attempts = int(row["total_login_attempts"])
            unique = int(row["unique_usernames"])
            score = float(row["bruteforce_score"])
            expected = round(attempts / max(1, unique), 4)
            assert score == pytest.approx(expected, rel=0, abs=1e-4), (
                f"bruteforce_score formula mismatch for {pcap_name}: expected {expected}, got {score}"
            )

    def test_protocol_category_exact_and_shared(self, comparison_sheet, expected_rows):
        actual = _comparison_by_pcap(comparison_sheet)
        categories = {str(row["protocol_category"]) for row in actual.values()}
        assert len(categories) == 1, f"Both rows must share the same protocol_category, got {categories}"

        expected_category = next(iter(expected_rows.values()))["protocol_category"]
        assert categories == {expected_category}, (
            f"protocol_category must exactly match the HTML-derived category {expected_category!r}, "
            f"got {categories}"
        )


class TestBehavioralComparison:
    def test_bruteforce_has_more_login_attempts(self, comparison_sheet):
        actual = _comparison_by_pcap(comparison_sheet)
        bf_attempts = int(actual["bruteforce.pcap"]["total_login_attempts"])
        ftp_attempts = int(actual["ftp.pcap"]["total_login_attempts"])
        assert bf_attempts > ftp_attempts, (
            f"Expected bruteforce.pcap to have more login attempts than ftp.pcap, "
            f"got {bf_attempts} vs {ftp_attempts}"
        )

    def test_bruteforce_has_higher_score(self, comparison_sheet):
        actual = _comparison_by_pcap(comparison_sheet)
        bf_score = float(actual["bruteforce.pcap"]["bruteforce_score"])
        ftp_score = float(actual["ftp.pcap"]["bruteforce_score"])
        assert bf_score > ftp_score, (
            f"Expected bruteforce.pcap to have a higher bruteforce_score than ftp.pcap, "
            f"got {bf_score} vs {ftp_score}"
        )

    def test_normal_ftp_has_success(self, comparison_sheet):
        actual = _comparison_by_pcap(comparison_sheet)
        ftp_success = int(actual["ftp.pcap"]["success_count"])
        assert ftp_success >= 1, (
            f"Expected ftp.pcap to have at least one successful login, got {ftp_success}"
        )


class TestMethodsSheet:
    def test_methods_sheet_not_empty(self, methods_sheet):
        non_empty = []
        for row in methods_sheet.iter_rows(values_only=True):
            if any(cell is not None and str(cell).strip() for cell in row):
                non_empty.append(row)
        assert non_empty, "methods sheet must contain text content"

    def test_methods_sheet_within_8_lines(self, methods_sheet):
        non_empty_rows = 0
        for row in methods_sheet.iter_rows(values_only=True):
            if any(cell is not None and str(cell).strip() for cell in row):
                non_empty_rows += 1
        assert non_empty_rows <= 8, (
            f"methods sheet must have no more than 8 non-empty rows, got {non_empty_rows}"
        )

    def test_methods_mentions_required_deterministic_rules(self, methods_sheet):
        text = []
        for row in methods_sheet.iter_rows(values_only=True):
            for cell in row:
                if cell is not None and str(cell).strip():
                    text.append(str(cell).lower())
        full_text = " ".join(text)

        assert ("220" in full_text or "banner" in full_text), (
            "methods sheet should describe how server vs client was identified using the 220 banner"
        )
        assert "pass" in full_text, (
            "methods sheet should describe how login attempts were counted from PASS commands"
        )
        assert "230" in full_text and "success" in full_text, (
            "methods sheet should describe how success was defined using response code 230"
        )


class TestExcelFormatting:
    def test_header_row_bold(self, comparison_sheet):
        for cell in comparison_sheet[1]:
            assert cell.font.bold, f"Header cell {cell.coordinate} in comparison sheet must be bold"

    def test_numeric_columns_right_aligned(self, comparison_sheet):
        header_index = {comparison_sheet.cell(row=1, column=i).value: i for i in range(1, comparison_sheet.max_column + 1)}
        numeric_indices = [header_index[col] for col in NUMERIC_COLUMNS]

        for row_idx in range(2, comparison_sheet.max_row + 1):
            row_has_data = any(
                comparison_sheet.cell(row=row_idx, column=i).value not in (None, "")
                for i in range(1, comparison_sheet.max_column + 1)
            )
            if not row_has_data:
                continue

            for col_idx in numeric_indices:
                cell = comparison_sheet.cell(row=row_idx, column=col_idx)
                assert cell.alignment.horizontal == "right", (
                    f"Numeric cell {cell.coordinate} must be right-aligned, got {cell.alignment.horizontal!r}"
                )

    def test_bruteforce_score_has_4_decimal_display(self, comparison_sheet):
        header_index = {comparison_sheet.cell(row=1, column=i).value: i for i in range(1, comparison_sheet.max_column + 1)}
        score_col = header_index["bruteforce_score"]

        for row_idx in range(2, comparison_sheet.max_row + 1):
            cell = comparison_sheet.cell(row=row_idx, column=score_col)
            if cell.value in (None, ""):
                continue
            fmt = str(cell.number_format or "")
            assert fmt != "General", (
                f"bruteforce_score cell {cell.coordinate} must have an explicit number format showing 4 decimals"
            )
            assert re.search(r"\.[#0]*0{4}", fmt), (
                f"bruteforce_score cell {cell.coordinate} must display 4 decimal places, got format {fmt!r}"
            )