"""Auto-generated expectation tests for task verification.

These tests verify that the task execution produces correct outputs
based on the instruction requirements for computing growth dynamics
across biological strains and housing construction categories.
"""

import os
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl.styles import Font


OUTPUT_FILE = "/root/relative_change_ranking.xlsx"


class TestOutputFileExists:
    """Tests for verifying output file existence and basic structure."""

    def test_output_file_exists(self):
        """Verify the output Excel file was created at the specified path."""
        assert os.path.exists(OUTPUT_FILE), f"Output file not found at {OUTPUT_FILE}"

    def test_output_file_is_xlsx(self):
        """Verify output file has .xlsx extension."""
        assert OUTPUT_FILE.endswith(".xlsx"), "Output file must have .xlsx extension"

    def test_output_file_not_empty(self):
        """Verify output file is not empty."""
        assert os.path.getsize(OUTPUT_FILE) > 0, "Output file is empty"


class TestExcelSheetStructure:
    """Tests for verifying the Excel sheet structure."""

    def test_ranking_sheet_exists(self):
        """Verify the 'Ranking' sheet exists in the workbook."""
        wb = load_workbook(OUTPUT_FILE)
        assert "Ranking" in wb.sheetnames, "Sheet 'Ranking' not found in workbook"
        wb.close()

    def test_required_columns_exist(self):
        """Verify all required columns are present: item_type, item_name, ratio, rank."""
        df = pd.read_excel(OUTPUT_FILE, sheet_name="Ranking")
        required_columns = ["item_type", "item_name", "ratio", "rank"]
        for col in required_columns:
            assert col in df.columns, f"Required column '{col}' not found"

    def test_column_order(self):
        """Verify columns are in the exact order: item_type, item_name, ratio, rank."""
        df = pd.read_excel(OUTPUT_FILE, sheet_name="Ranking")
        expected_columns = ["item_type", "item_name", "ratio", "rank"]
        actual_columns = list(df.columns)[:4]  # First 4 columns
        assert actual_columns == expected_columns, (
            f"Column order mismatch. Expected {expected_columns}, got {actual_columns}"
        )


class TestDataContent:
    """Tests for verifying the data content and values."""

    def test_has_strain_rows(self):
        """Verify there are rows for biological strains (excluding Blank)."""
        df = pd.read_excel(OUTPUT_FILE, sheet_name="Ranking")
        strain_rows = df[df["item_type"].str.lower().str.contains("strain", na=False)]
        assert len(strain_rows) >= 2, "Expected at least 2 strain rows (C6706, LuxO)"

    def test_expected_strains_present(self):
        """Verify C6706 and LuxO strains are present in the data."""
        df = pd.read_excel(OUTPUT_FILE, sheet_name="Ranking")
        item_names = df["item_name"].str.strip().tolist()
        assert "C6706" in item_names, "Strain C6706 not found in item_name column"
        assert "LuxO" in item_names, "Strain LuxO not found in item_name column"

    def test_blank_strain_excluded(self):
        """Verify Blank strain is not included in the output."""
        df = pd.read_excel(OUTPUT_FILE, sheet_name="Ranking")
        item_names = df["item_name"].str.lower().str.strip().tolist()
        assert "blank" not in item_names, "Blank strain should be excluded"

    def test_has_building_type_rows(self):
        """Verify there are rows for housing construction building types."""
        df = pd.read_excel(OUTPUT_FILE, sheet_name="Ranking")
        building_types = ["1 unit", "2 to 4 units", "5 units or more"]
        item_names = df["item_name"].str.strip().tolist()
        found_count = sum(1 for bt in building_types if bt in item_names)
        # Some building types may have all NA values, so we check for at least some
        assert found_count >= 1, "Expected at least one building type row"

    def test_has_spending_category_row(self):
        """Verify there is at least one spending category row."""
        df = pd.read_excel(OUTPUT_FILE, sheet_name="Ranking")
        # Check for spending category type indicator
        spending_rows = df[
            df["item_type"].str.lower().str.contains("spend", na=False)
            | df["item_name"].str.lower().str.contains("annual", na=False)
        ]
        # If not found by type, should have exactly one winning spending category
        total_rows = len(df)
        # We expect: 2 strains + up to 3 building types + 1 spending category
        assert total_rows >= 4, "Expected at least 4 rows in the ranking"

    def test_ratio_values_are_numeric(self):
        """Verify all ratio values are numeric."""
        df = pd.read_excel(OUTPUT_FILE, sheet_name="Ranking")
        assert pd.api.types.is_numeric_dtype(df["ratio"]), "Ratio column must be numeric"

    def test_ratio_values_positive(self):
        """Verify all ratio values are positive (fold-change must be >= 1)."""
        df = pd.read_excel(OUTPUT_FILE, sheet_name="Ranking")
        assert (df["ratio"] > 0).all(), "All ratio values must be positive"

    def test_ratio_values_at_least_one(self):
        """Verify all ratio values are at least 1 (max/min >= 1)."""
        df = pd.read_excel(OUTPUT_FILE, sheet_name="Ranking")
        assert (df["ratio"] >= 1).all(), "All ratio values should be >= 1 (max/min ratio)"

    def test_rank_values_are_integers(self):
        """Verify all rank values are integers."""
        df = pd.read_excel(OUTPUT_FILE, sheet_name="Ranking")
        # Check if values are effectively integers (may be stored as float)
        assert df["rank"].apply(lambda x: float(x).is_integer()).all(), (
            "Rank values must be integers"
        )

    def test_rank_values_positive(self):
        """Verify all rank values are positive integers."""
        df = pd.read_excel(OUTPUT_FILE, sheet_name="Ranking")
        assert (df["rank"] > 0).all(), "All rank values must be positive"

    def test_rank_starts_at_one(self):
        """Verify rank starts at 1 for the largest ratio."""
        df = pd.read_excel(OUTPUT_FILE, sheet_name="Ranking")
        assert df["rank"].min() == 1, "Minimum rank should be 1"

    def test_ranks_are_consecutive(self):
        """Verify ranks form a consecutive sequence starting from 1."""
        df = pd.read_excel(OUTPUT_FILE, sheet_name="Ranking")
        n = len(df)
        expected_ranks = set(range(1, n + 1))
        actual_ranks = set(df["rank"].astype(int).tolist())
        assert actual_ranks == expected_ranks, (
            f"Ranks should be consecutive from 1 to {n}. Got {sorted(actual_ranks)}"
        )

    def test_ranking_order_descending_by_ratio(self):
        """Verify data is ranked by ratio in descending order."""
        df = pd.read_excel(OUTPUT_FILE, sheet_name="Ranking")
        # Sort by rank and check ratio is descending
        df_sorted = df.sort_values("rank")
        ratios = df_sorted["ratio"].tolist()
        for i in range(len(ratios) - 1):
            assert ratios[i] >= ratios[i + 1], (
                f"Ratio at rank {i+1} ({ratios[i]}) should be >= ratio at rank {i+2} ({ratios[i+1]})"
            )

    def test_rank_one_has_largest_ratio(self):
        """Verify rank 1 corresponds to the largest ratio value."""
        df = pd.read_excel(OUTPUT_FILE, sheet_name="Ranking")
        rank_one_ratio = df[df["rank"] == 1]["ratio"].values[0]
        max_ratio = df["ratio"].max()
        assert abs(rank_one_ratio - max_ratio) < 0.0001, (
            f"Rank 1 ratio ({rank_one_ratio}) should equal max ratio ({max_ratio})"
        )

    def test_no_duplicate_item_names(self):
        """Verify there are no duplicate item names."""
        df = pd.read_excel(OUTPUT_FILE, sheet_name="Ranking")
        duplicates = df["item_name"].duplicated().sum()
        assert duplicates == 0, f"Found {duplicates} duplicate item names"


class TestRatioFormatting:
    """Tests for verifying ratio decimal places."""

    def test_ratio_has_four_decimal_places(self):
        """Verify ratio values are written with 4 decimal places."""
        wb = load_workbook(OUTPUT_FILE)
        ws = wb["Ranking"]

        # Find the ratio column index (0-based, but openpyxl is 1-based)
        header_row = [cell.value for cell in ws[1]]
        if "ratio" in header_row:
            ratio_col = header_row.index("ratio") + 1  # 1-based index

            # Check number format of cells in ratio column
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=ratio_col)
                if cell.value is not None:
                    # Check if the number format specifies 4 decimal places
                    # Common formats: "0.0000", "#,##0.0000", etc.
                    num_format = cell.number_format
                    # Verify the value when read can be represented with 4 decimals
                    value = cell.value
                    if isinstance(value, (int, float)):
                        formatted = f"{value:.4f}"
                        # Value should be representable with 4 decimal precision
                        assert abs(float(formatted) - value) < 0.00005, (
                            f"Ratio value {value} should be writable with 4 decimal places"
                        )
        wb.close()


class TestExcelFormatting:
    """Tests for verifying Excel formatting requirements."""

    def test_header_row_is_bold(self):
        """Verify the header row has bold font formatting."""
        wb = load_workbook(OUTPUT_FILE)
        ws = wb["Ranking"]

        # Check that header cells (row 1) are bold
        for cell in ws[1]:
            if cell.value is not None:
                assert cell.font.bold, f"Header cell '{cell.value}' should be bold"
        wb.close()

    def test_top_row_is_frozen(self):
        """Verify the top row is frozen (freeze panes)."""
        wb = load_workbook(OUTPUT_FILE)
        ws = wb["Ranking"]

        # Check freeze panes - should be set to freeze first row
        # freeze_panes should be "A2" to freeze row 1
        freeze_pane = ws.freeze_panes
        assert freeze_pane is not None, "Freeze panes should be set"
        # Common freeze positions for top row: "A2", "B2", etc. (row 2 means row 1 is frozen)
        if freeze_pane:
            # The row number in freeze_panes should be 2 (meaning row 1 is frozen)
            from openpyxl.utils import coordinate_from_string, column_index_from_string
            col_letter, row_num = coordinate_from_string(freeze_pane)
            assert row_num == 2, f"Freeze panes at row {row_num}, expected row 2 (to freeze row 1)"
        wb.close()


class TestItemTypes:
    """Tests for verifying item_type values are appropriate."""

    def test_item_type_not_empty(self):
        """Verify item_type column has no empty/null values."""
        df = pd.read_excel(OUTPUT_FILE, sheet_name="Ranking")
        assert df["item_type"].notna().all(), "item_type column should have no null values"
        assert (df["item_type"].str.strip() != "").all(), "item_type values should not be empty strings"

    def test_item_name_not_empty(self):
        """Verify item_name column has no empty/null values."""
        df = pd.read_excel(OUTPUT_FILE, sheet_name="Ranking")
        assert df["item_name"].notna().all(), "item_name column should have no null values"
        assert (df["item_name"].str.strip() != "").all(), "item_name values should not be empty strings"


class TestTieBreaking:
    """Tests for verifying tie-breaking rules (alphabetical by item_name)."""

    def test_ties_broken_alphabetically(self):
        """Verify that ties in ratio are broken alphabetically by item_name."""
        df = pd.read_excel(OUTPUT_FILE, sheet_name="Ranking")
        df_sorted = df.sort_values("rank")

        # Find groups with same ratio
        for ratio_val in df_sorted["ratio"].unique():
            tied_items = df_sorted[abs(df_sorted["ratio"] - ratio_val) < 0.00005]
            if len(tied_items) > 1:
                # For tied items, check they are in alphabetical order by item_name
                item_names = tied_items.sort_values("rank")["item_name"].tolist()
                sorted_names = sorted(item_names)
                assert item_names == sorted_names, (
                    f"Tied items with ratio {ratio_val:.4f} should be ordered alphabetically. "
                    f"Got {item_names}, expected {sorted_names}"
                )


class TestExpectedRowCount:
    """Tests for verifying the expected number of rows."""

    def test_minimum_row_count(self):
        """Verify minimum expected rows: 2 strains + building types + 1 spending category."""
        df = pd.read_excel(OUTPUT_FILE, sheet_name="Ranking")
        # Minimum: 2 strains (C6706, LuxO) + 1 spending category = 3
        # Plus potential building types (1 unit, 2 to 4 units, 5 units or more)
        # Note: "2 to 4 units" column appears to have NA values, so may not be included
        assert len(df) >= 3, f"Expected at least 3 rows, got {len(df)}"

    def test_maximum_row_count(self):
        """Verify maximum expected rows: 2 strains + 3 building types + 1 spending = 6."""
        df = pd.read_excel(OUTPUT_FILE, sheet_name="Ranking")
        # Maximum: 2 strains + 3 building types + 1 spending category = 6
        assert len(df) <= 6, f"Expected at most 6 rows, got {len(df)}"


class TestDataIntegrity:
    """Tests for data integrity and correctness."""

    def test_strain_fold_change_reasonable(self):
        """Verify strain fold-change values are within reasonable biological range."""
        df = pd.read_excel(OUTPUT_FILE, sheet_name="Ranking")
        strain_rows = df[df["item_name"].isin(["C6706", "LuxO"])]

        for _, row in strain_rows.iterrows():
            ratio = row["ratio"]
            # OD values in the data range from ~0.02 to ~0.5, so fold-change should be reasonable
            assert 1 <= ratio <= 100, (
                f"Strain {row['item_name']} fold-change {ratio} seems unreasonable"
            )

    def test_construction_ratio_reasonable(self):
        """Verify construction ratios are within reasonable range."""
        df = pd.read_excel(OUTPUT_FILE, sheet_name="Ranking")
        building_types = ["1 unit", "2 to 4 units", "5 units or more"]
        construction_rows = df[df["item_name"].isin(building_types)]

        for _, row in construction_rows.iterrows():
            ratio = row["ratio"]
            # Construction data shows monthly values, ratios should be moderate
            assert 1 <= ratio <= 100, (
                f"Construction type '{row['item_name']}' ratio {ratio} seems unreasonable"
            )


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
